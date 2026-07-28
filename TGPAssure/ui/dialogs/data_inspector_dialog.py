from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

import numpy as np
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QFormLayout,
    QLabel,
    QPlainTextEdit,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
except Exception:  # pragma: no cover - optional runtime fallback
    FigureCanvas = None
    Figure = None


class DataInspectorDialog(QDialog):
    """Immediate statistics, histogram and cross-plot inspection for tabular data.

    The generic inspector deliberately samples very large files instead of loading
    an unlimited table into GUI memory. Module-specific geophysical viewers remain
    the authoritative tools for full-resolution interpretation/export.
    """

    MAX_ROWS = 200_000
    MAX_SCATTER_POINTS = 20_000

    def __init__(self, path: str | Path, parent=None, initial_tab: str = "Statistics") -> None:
        super().__init__(parent)
        self.path = Path(path).expanduser().resolve()
        if not self.path.is_file():
            raise FileNotFoundError(self.path)
        self.setWindowTitle(f"Data Inspector — {self.path.name}")
        self.resize(900, 650)
        self.setMinimumSize(680, 480)

        self._headers, self._rows, self._truncated = self._load(self.path)
        self._numeric = self._numeric_columns()

        root = QVBoxLayout(self)
        status = QLabel(self._status_text(), self)
        status.setWordWrap(True)
        status.setTextInteractionFlags(Qt.TextSelectableByMouse)
        root.addWidget(status)

        self.tabs = QTabWidget(self)
        root.addWidget(self.tabs, 1)
        self._build_statistics()
        self._build_histogram()
        self._build_crossplot()

        for index in range(self.tabs.count()):
            if self.tabs.tabText(index).lower() == initial_tab.lower():
                self.tabs.setCurrentIndex(index)
                break

    def _status_text(self) -> str:
        size_mb = self.path.stat().st_size / (1024 * 1024)
        sample_note = f" — first {len(self._rows):,} rows sampled" if self._truncated else ""
        return (
            f"{self.path.name} · {size_mb:.2f} MB · {len(self._headers):,} columns · "
            f"{len(self._rows):,} data rows loaded{sample_note}"
        )

    @classmethod
    def _load(cls, path: Path) -> tuple[list[str], list[list[object]], bool]:
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                iterator = sheet.iter_rows(values_only=True)
                header_row = next(iterator, None)
                headers = [str(value or "") for value in header_row] if header_row else []
                rows: list[list[object]] = []
                truncated = False
                for row_index, row in enumerate(iterator):
                    if row_index >= cls.MAX_ROWS:
                        truncated = True
                        break
                    rows.append(list(row))
                return headers, rows, truncated
            finally:
                workbook.close()

        sample = path.read_text(encoding="utf-8", errors="replace")[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample.strip() else csv.excel
        except csv.Error:
            dialect = csv.excel

        with path.open("r", encoding="utf-8", errors="replace", newline="") as stream:
            reader = csv.reader(stream, dialect)
            headers = next(reader, [])
            rows = []
            truncated = False
            for row_index, row in enumerate(reader):
                if row_index >= cls.MAX_ROWS:
                    truncated = True
                    break
                rows.append(row)
        return [str(value or "") for value in headers], rows, truncated

    def _numeric_columns(self) -> dict[str, np.ndarray]:
        result: dict[str, np.ndarray] = {}
        row_count = len(self._rows)
        for column_index, raw_name in enumerate(self._headers):
            values = np.full(row_count, np.nan, dtype=float)
            finite_count = 0
            for row_index, row in enumerate(self._rows):
                if column_index >= len(row):
                    continue
                try:
                    value = float(row[column_index])
                except (TypeError, ValueError):
                    continue
                if np.isfinite(value):
                    values[row_index] = value
                    finite_count += 1
            if finite_count:
                name = raw_name.strip() or f"Column {column_index + 1}"
                # Preserve duplicate headers without silently overwriting a column.
                unique_name = name
                suffix = 2
                while unique_name in result:
                    unique_name = f"{name} ({suffix})"
                    suffix += 1
                result[unique_name] = values
        return result

    def _build_statistics(self) -> None:
        page = QWidget(self)
        layout = QVBoxLayout(page)
        text = QPlainTextEdit(page)
        text.setReadOnly(True)
        lines = [
            f"File: {self.path}",
            f"Rows loaded: {len(self._rows):,}" + (" (sampled)" if self._truncated else ""),
            f"Columns: {len(self._headers):,}",
            f"Numeric columns: {len(self._numeric):,}",
            "",
        ]
        if not self._numeric:
            lines.append("No numeric columns were detected in the loaded rows.")
        for name, raw_values in self._numeric.items():
            values = raw_values[np.isfinite(raw_values)]
            if not values.size:
                continue
            lines.extend(
                [
                    name,
                    f"  count = {values.size:,}",
                    f"  missing/non-numeric = {raw_values.size - values.size:,}",
                    f"  min = {np.min(values):.8g}",
                    f"  max = {np.max(values):.8g}",
                    f"  mean = {np.mean(values):.8g}",
                    f"  median = {np.median(values):.8g}",
                    f"  std = {np.std(values):.8g}",
                    f"  p05 / p95 = {np.percentile(values, 5):.8g} / {np.percentile(values, 95):.8g}",
                    "",
                ]
            )
        text.setPlainText("\n".join(lines))
        layout.addWidget(text)
        self.tabs.addTab(page, "Statistics")

    def _build_histogram(self) -> None:
        page = QWidget(self)
        layout = QVBoxLayout(page)
        combo = QComboBox(page)
        combo.addItems(self._numeric.keys())
        layout.addWidget(combo)

        if FigureCanvas is not None and Figure is not None:
            figure = Figure(tight_layout=True)
            canvas = FigureCanvas(figure)
            layout.addWidget(canvas, 1)

            def refresh() -> None:
                figure.clear()
                axis = figure.add_subplot(111)
                raw = self._numeric.get(combo.currentText())
                if raw is not None:
                    values = raw[np.isfinite(raw)]
                    if values.size:
                        bins = min(80, max(10, int(np.sqrt(values.size))))
                        axis.hist(values, bins=bins)
                        axis.set_title(combo.currentText())
                        axis.set_xlabel("Value")
                        axis.set_ylabel("Frequency")
                canvas.draw_idle()
        else:
            output = QPlainTextEdit(page)
            output.setReadOnly(True)
            layout.addWidget(output, 1)

            def refresh() -> None:
                raw = self._numeric.get(combo.currentText())
                if raw is None:
                    output.clear()
                    return
                values = raw[np.isfinite(raw)]
                if not values.size:
                    output.setPlainText("No finite numeric samples.")
                    return
                counts, edges = np.histogram(values, bins=min(30, max(5, int(np.sqrt(values.size)))))
                output.setPlainText(
                    "\n".join(
                        f"{edges[index]:.6g} – {edges[index + 1]:.6g}: {counts[index]}"
                        for index in range(len(counts))
                    )
                )

        combo.currentTextChanged.connect(refresh)
        refresh()
        self.tabs.addTab(page, "Histogram")

    def _build_crossplot(self) -> None:
        page = QWidget(self)
        layout = QVBoxLayout(page)
        form = QFormLayout()
        x_combo = QComboBox(page)
        y_combo = QComboBox(page)
        x_combo.addItems(self._numeric.keys())
        y_combo.addItems(self._numeric.keys())
        if y_combo.count() > 1:
            y_combo.setCurrentIndex(1)
        form.addRow("X", x_combo)
        form.addRow("Y", y_combo)
        layout.addLayout(form)

        summary = QLabel(page)
        summary.setWordWrap(True)
        summary.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(summary)

        figure = Figure(tight_layout=True) if Figure is not None else None
        canvas = FigureCanvas(figure) if figure is not None and FigureCanvas is not None else None
        if canvas is not None:
            layout.addWidget(canvas, 1)
        else:
            fallback = QPlainTextEdit(page)
            fallback.setReadOnly(True)
            layout.addWidget(fallback, 1)

        def refresh() -> None:
            x_raw = self._numeric.get(x_combo.currentText())
            y_raw = self._numeric.get(y_combo.currentText())
            if x_raw is None or y_raw is None:
                summary.setText("Select two numeric columns.")
                return
            mask = np.isfinite(x_raw) & np.isfinite(y_raw)
            x_values = x_raw[mask]
            y_values = y_raw[mask]
            count = x_values.size
            correlation = float(np.corrcoef(x_values, y_values)[0, 1]) if count > 1 else float("nan")
            summary.setText(f"Paired finite samples: {count:,} · Pearson correlation: {correlation:.6g}")

            if canvas is not None and figure is not None:
                figure.clear()
                axis = figure.add_subplot(111)
                if count:
                    if count > self.MAX_SCATTER_POINTS:
                        indices = np.linspace(0, count - 1, self.MAX_SCATTER_POINTS, dtype=int)
                        plot_x, plot_y = x_values[indices], y_values[indices]
                    else:
                        plot_x, plot_y = x_values, y_values
                    axis.scatter(plot_x, plot_y, s=8, alpha=0.55)
                axis.set_xlabel(x_combo.currentText())
                axis.set_ylabel(y_combo.currentText())
                axis.set_title("Cross Plot")
                canvas.draw_idle()
            else:
                fallback.setPlainText(
                    "Matplotlib is unavailable. Numeric correlation is shown above; "
                    "install the declared matplotlib dependency for graphical cross-plots."
                )

        x_combo.currentTextChanged.connect(refresh)
        y_combo.currentTextChanged.connect(refresh)
        refresh()
        self.tabs.addTab(page, "Cross Plot")
