from __future__ import annotations

import pytest
import tempfile
import shutil
import json
from pathlib import Path

from PySide6.QtCore import Qt, QCoreApplication
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QApplication
from PySide6.QtTest import QTest

from core.data_access.db_engine import DatabaseEngine
from core.data_access.project_repository import ProjectRepository
from core.infrastructure.service_container import ServiceContainer
from core.infrastructure.job_manager import JobManager
from core.domain.qc_engine import QCStatus
from modules.seismic.segy_qc.segy_qc_view import SegyQcView
from modules.seismic.segy_qc.segy_qc_controller import SegyQcController
from report.report_model import ReportModel, TextSection, TableSection
from report.renderers.pdf_renderer import PdfRenderer
from report.renderers.xlsx_renderer import XlsxRenderer


@pytest.fixture
def app() -> QApplication:
    return QApplication.instance() or QApplication([])


@pytest.fixture
def temp_dir() -> Path:
    path = Path(tempfile.mkdtemp())
    yield path
    shutil.rmtree(path, ignore_errors=True)


@pytest.fixture
def db_engine(temp_dir: Path) -> DatabaseEngine:
    db_path = temp_dir / "test.db"
    engine = DatabaseEngine(db_path)
    conn = engine.get_write_connection()
    try:
        conn.execute("""
            CREATE TABLE project (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                project_uuid TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                schema_version INTEGER NOT NULL DEFAULT 0
            )
        """)
        conn.execute("INSERT INTO project (id, project_uuid, name) VALUES (1, 'test', 'Test')")
        conn.execute("""
            CREATE TABLE project_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL DEFAULT 1,
                file_uuid TEXT NOT NULL UNIQUE,
                module TEXT NOT NULL,
                file_role TEXT NOT NULL,
                original_name TEXT NOT NULL,
                display_name TEXT NOT NULL,
                absolute_path TEXT,
                relative_path TEXT,
                extension TEXT,
                mime_type TEXT,
                size_bytes INTEGER NOT NULL DEFAULT 0,
                sha256 TEXT,
                status TEXT NOT NULL DEFAULT 'available',
                metadata_json TEXT NOT NULL DEFAULT '{}',
                imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                last_verified_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL DEFAULT 1,
                file_id INTEGER,
                job_uuid TEXT NOT NULL UNIQUE,
                job_type TEXT NOT NULL,
                module TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'queued',
                priority INTEGER NOT NULL DEFAULT 0,
                progress REAL NOT NULL DEFAULT 0.0,
                message TEXT,
                payload_json TEXT NOT NULL DEFAULT '{}',
                result_json TEXT,
                error_text TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                started_at TEXT,
                finished_at TEXT,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE qc_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL DEFAULT 1,
                file_id INTEGER,
                job_id INTEGER,
                run_uuid TEXT NOT NULL UNIQUE,
                module TEXT NOT NULL,
                qc_profile TEXT NOT NULL,
                profile_version TEXT,
                status TEXT NOT NULL DEFAULT 'pending',
                overall_result TEXT NOT NULL DEFAULT 'pending',
                score REAL,
                parameters_json TEXT NOT NULL DEFAULT '{}',
                summary_json TEXT NOT NULL DEFAULT '{}',
                started_at TEXT,
                completed_at TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("""
            CREATE TABLE qc_stage_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                qc_run_id INTEGER NOT NULL,
                stage_key TEXT NOT NULL,
                stage_name TEXT NOT NULL,
                stage_order INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                result TEXT NOT NULL DEFAULT 'pending',
                score REAL,
                metrics_json TEXT NOT NULL DEFAULT '{}',
                message TEXT,
                started_at TEXT,
                completed_at TEXT,
                duration_ms INTEGER
            )
        """)
        conn.execute("""
            CREATE TABLE qc_findings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                qc_run_id INTEGER NOT NULL,
                stage_result_id INTEGER,
                file_id INTEGER,
                finding_code TEXT NOT NULL,
                severity TEXT NOT NULL,
                category TEXT NOT NULL,
                title TEXT NOT NULL,
                description TEXT NOT NULL,
                metric_name TEXT,
                observed_value REAL,
                expected_min REAL,
                expected_max REAL,
                unit TEXT,
                station_id TEXT,
                line_id TEXT,
                sample_index INTEGER,
                timestamp_utc TEXT,
                location_x REAL,
                location_y REAL,
                location_z REAL,
                crs TEXT,
                context_json TEXT NOT NULL DEFAULT '{}',
                is_resolved INTEGER NOT NULL DEFAULT 0,
                resolution_note TEXT,
                resolved_at TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
    finally:
        conn.close()
    return engine


@pytest.fixture
def container(app: QApplication, db_engine: DatabaseEngine) -> ServiceContainer:
    container = ServiceContainer()
    container.register(DatabaseEngine, db_engine)
    job_manager = JobManager(db_engine)
    container.register(JobManager, job_manager)
    return container


@pytest.fixture
def controller(app: QApplication, db_engine: DatabaseEngine, container: ServiceContainer) -> SegyQcController:
    project_repo = ProjectRepository(db_engine)
    job_manager = container.resolve(JobManager)
    return SegyQcController(db_engine, job_manager, project_repo)


def test_stage_list_colors_after_run(qtbot, app: QApplication, controller: SegyQcController) -> None:
    view = SegyQcView(controller)
    view.show()
    qtbot.addWidget(view)
    
    view.set_stages(["stage1", "stage2", "stage3"])
    
    item1 = view._stage_items.get("stage1")
    item2 = view._stage_items.get("stage2")
    item3 = view._stage_items.get("stage3")
    
    assert item1 is not None
    assert item2 is not None
    assert item3 is not None
    
    item1.set_status(QCStatus.PASS)
    item2.set_status(QCStatus.WARN)
    item3.set_status(QCStatus.FAIL)
    
    assert item1.background().color() == QColor(0, 176, 80)
    assert item2.background().color() == QColor(255, 200, 0)
    assert item3.background().color() == QColor(255, 0, 0)


def test_pdf_renderer_creates_valid_pdf(temp_dir: Path) -> None:
    model = ReportModel("Test Report")
    model.add_section(TextSection("Summary", "This is a test summary."))
    model.add_section(TableSection(
        "Test Table",
        ["Column 1", "Column 2"],
        [["Value 1", "Value 2"], ["Value 3", "Value 4"]]
    ))
    
    output_path = temp_dir / "test_report.pdf"
    renderer = PdfRenderer()
    result = renderer.render(model, output_path)
    
    assert result.exists()
    assert result.stat().st_size > 0
    assert result.suffix == ".pdf"


def test_xlsx_renderer_creates_valid_xlsx(temp_dir: Path) -> None:
    model = ReportModel("Test Report")
    model.add_section(TableSection(
        "Test Table",
        ["Column 1", "Column 2"],
        [["Value 1", "Value 2"], ["Value 3", "Value 4"]]
    ))
    
    output_path = temp_dir / "test_report.xlsx"
    renderer = XlsxRenderer()
    result = renderer.render(model, output_path)
    
    assert result.exists()
    assert result.stat().st_size > 0
    assert result.suffix == ".xlsx"


def test_xlsx_renderer_has_correct_row_count(temp_dir: Path) -> None:
    model = ReportModel("Test Report")
    model.add_section(TableSection(
        "Test Table",
        ["Col A", "Col B"],
        [["Row1A", "Row1B"], ["Row2A", "Row2B"], ["Row3A", "Row3B"]]
    ))
    
    output_path = temp_dir / "test_report.xlsx"
    renderer = XlsxRenderer()
    result = renderer.render(model, output_path)
    
    from openpyxl import load_workbook
    wb = load_workbook(result)
    ws = wb.active
    
    assert ws.max_row >= 5


def test_segy_qc_view_initial_state(controller: SegyQcController) -> None:
    view = SegyQcView(controller)
    
    assert view.run_button.isEnabled() is True
    assert view.report_button.isEnabled() is False
    assert view.status_label.text() == "Ready"


def test_segy_qc_view_reset(controller: SegyQcController) -> None:
    view = SegyQcView(controller)
    view.set_stages(["stage1", "stage2"])
    
    view.reset_view()
    
    for section in view._stage_sections.values():
        assert section.status_label.text() == "Pending"
    
    assert view.report_button.isEnabled() is False
    assert view.run_button.isEnabled() is True