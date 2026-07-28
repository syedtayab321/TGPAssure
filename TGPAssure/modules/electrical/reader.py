from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

import numpy as np

from modules.electrical.constants import COLUMN_ALIASES, ElectricalMethod, SUPPORTED_EXTENSIONS
from modules.electrical.models import ElectricalDataset


_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ρ", "rho").replace("φ", "phi")
    text = text.replace("ω", "ohm").replace("Ω", "ohm")
    return _NORMALIZE_RE.sub("_", text).strip("_")


_ALIAS_LOOKUP: dict[str, str] = {}
for canonical, aliases in COLUMN_ALIASES.items():
    _ALIAS_LOOKUP[normalize_header(canonical)] = canonical
    for alias in aliases:
        _ALIAS_LOOKUP[normalize_header(alias)] = canonical


class ElectricalReader:
    """Reader for QC-oriented electrical exports.

    Supports CSV/TXT/DAT/XYZ/TSV and XLSX/XLSM. It intentionally targets
    documented/exported tables rather than pretending to decode proprietary
    binary instrument formats without a published specification.
    """

    def inspect(self, path: str | Path) -> dict[str, Any]:
        source = Path(path).expanduser().resolve()
        self._validate_path(source)
        headers, preview, source_format = self._read_preview(source)
        mapping = self._map_headers(headers)
        mapped_fields = set(mapping.values())
        method = self.infer_method(mapped_fields)
        measurement_fields = {
            "current_ma", "voltage_mv", "resistance_ohm", "apparent_resistivity_ohm_m",
            "chargeability_mv_v", "sp_mv", "frequency_hz", "phase_mrad", "phase_deg",
            "electric_field_mv_km", "electric_field_x_mv_km", "electric_field_y_mv_km",
        }
        is_candidate = bool(mapped_fields & measurement_fields)
        return {
            "source_path": str(source),
            "source_file": source.name,
            "size_bytes": source.stat().st_size,
            "source_format": source_format,
            "headers": headers,
            "mapped_fields": mapping,
            "preview_rows": preview,
            "inferred_method": method.value,
            "is_electrical_candidate": is_candidate,
        }

    def read(self, path: str | Path, method: ElectricalMethod | str = ElectricalMethod.AUTO) -> ElectricalDataset:
        source = Path(path).expanduser().resolve()
        self._validate_path(source)
        headers, rows, source_format = self._read_all(source)
        if not rows:
            raise ValueError("The selected electrical file contains no data rows.")
        mapping = self._map_headers(headers)
        if not mapping:
            raise ValueError(
                "No recognized electrical columns were found. Export a tabular file containing fields such as "
                "A/B/M/N, current, voltage, apparent resistivity, chargeability, SP, frequency or phase."
            )

        canonical_columns: dict[str, list[Any]] = {canonical: [] for canonical in set(mapping.values())}
        for row in rows:
            for raw_header, canonical in mapping.items():
                canonical_columns[canonical].append(row.get(raw_header))

        columns: dict[str, np.ndarray] = {}
        text_fields = {"line_id", "timestamp", "source_id", "repeat_id", "array_type"}
        bool_fields = {"is_base"}
        for name, values in canonical_columns.items():
            if name in text_fields:
                columns[name] = np.asarray(["" if v is None else str(v).strip() for v in values], dtype=object)
            elif name in bool_fields:
                columns[name] = np.asarray([_to_bool(v) for v in values], dtype=bool)
            else:
                columns[name] = np.asarray([_to_float(v) for v in values], dtype=float)

        selected = ElectricalMethod(str(method)) if not isinstance(method, ElectricalMethod) else method
        inferred = self.infer_method(set(columns))
        if selected == ElectricalMethod.AUTO:
            selected = inferred

        dataset = ElectricalDataset(
            source_path=source,
            method=selected,
            columns=columns,
            raw_headers=headers,
            metadata={
                "mapped_columns": {raw: canonical for raw, canonical in mapping.items()},
                "inferred_method": inferred.value,
                "size_bytes": source.stat().st_size,
            },
            source_format=source_format,
        )
        self._normalize_units(dataset, headers, mapping)
        return dataset

    @staticmethod
    def infer_method(fields: set[str]) -> ElectricalMethod:
        if "frequency_hz" in fields and ("phase_mrad" in fields or "phase_deg" in fields):
            return ElectricalMethod.SIP
        if "frequency_hz" in fields and "chargeability_mv_v" in fields:
            return ElectricalMethod.FDIP
        if "source_id" in fields and ("voltage_mv" in fields or "sp_mv" in fields):
            return ElectricalMethod.MALM
        if "sp_mv" in fields and not ({"current_ma", "a", "b"} & fields):
            return ElectricalMethod.SP
        # Sounding geometry is more specific than the presence of an optional
        # chargeability column: many VES field sheets contain both rho-a and IP.
        if "ab2_m" in fields:
            return ElectricalMethod.VES
        if "chargeability_mv_v" in fields:
            return ElectricalMethod.TDIP
        if {"a", "b", "m", "n"}.issubset(fields):
            return ElectricalMethod.ERT
        return ElectricalMethod.PROFILING

    def _read_preview(self, path: Path) -> tuple[list[str], list[dict[str, Any]], str]:
        headers, rows, source_format = self._read_all(path, max_rows=12)
        return headers, rows, source_format

    def _read_all(self, path: Path, max_rows: int | None = None) -> tuple[list[str], list[dict[str, Any]], str]:
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return self._read_excel(path, max_rows)
        return self._read_delimited(path, max_rows)

    def _read_excel(self, path: Path, max_rows: int | None) -> tuple[list[str], list[dict[str, Any]], str]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration:
                return [], [], "xlsx"
            headers = [str(value).strip() if value is not None else f"column_{i + 1}" for i, value in enumerate(header_row)]
            rows: list[dict[str, Any]] = []
            for index, values in enumerate(iterator):
                if max_rows is not None and index >= max_rows:
                    break
                if not any(value not in (None, "") for value in values):
                    continue
                rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
            return headers, rows, "xlsx"
        finally:
            workbook.close()

    def _read_delimited(self, path: Path, max_rows: int | None) -> tuple[list[str], list[dict[str, Any]], str]:
        sample = _read_text_prefix(path, 65536)
        if not sample.strip():
            return [], [], "delimited"
        delimiter = _detect_delimiter(sample)
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
            if delimiter is None:
                return self._read_whitespace(stream, max_rows)
            reader = csv.DictReader(stream, delimiter=delimiter)
            headers = [str(value or "").strip() for value in (reader.fieldnames or [])]
            rows: list[dict[str, Any]] = []
            for index, row in enumerate(reader):
                if max_rows is not None and index >= max_rows:
                    break
                if not row or not any(str(v or "").strip() for v in row.values()):
                    continue
                rows.append({str(key or "").strip(): value for key, value in row.items()})
            return headers, rows, "delimited"

    def _read_whitespace(self, stream: io.TextIOBase, max_rows: int | None) -> tuple[list[str], list[dict[str, Any]], str]:
        header_line: str | None = None
        for raw_line in stream:
            line = raw_line.strip()
            if line and not line.lstrip().startswith(("#", ";", "//")):
                header_line = line
                break
        if header_line is None:
            return [], [], "whitespace"
        headers = re.split(r"\s+", header_line)
        rows: list[dict[str, Any]] = []
        for raw_line in stream:
            line = raw_line.strip()
            if not line or line.lstrip().startswith(("#", ";", "//")):
                continue
            values = re.split(r"\s+", line)
            if len(values) < 2:
                continue
            rows.append({header: values[i] if i < len(values) else None for i, header in enumerate(headers)})
            if max_rows is not None and len(rows) >= max_rows:
                break
        return headers, rows, "whitespace"

    @staticmethod
    def _map_headers(headers: list[str]) -> dict[str, str]:
        mapping: dict[str, str] = {}
        used: set[str] = set()
        normalized_headers = {normalize_header(header) for header in headers}
        has_explicit_abmn_geometry = bool({"a", "b", "n"}.intersection(normalized_headers))
        for header in headers:
            normalized = normalize_header(header)
            # A bare M is ambiguous in electrical exports: it is the potential
            # electrode in ABMN tables, but often means chargeability in VES/IP
            # field sheets. Resolve it from the surrounding schema.
            if normalized == "m":
                canonical = "m" if has_explicit_abmn_geometry else "chargeability_mv_v"
            elif re.fullmatch(r"m\d+", normalized):
                canonical = f"window_{int(normalized[1:]):02d}"
            elif normalized.startswith("decay_") or normalized.startswith("window_"):
                canonical = normalized
            else:
                canonical = _ALIAS_LOOKUP.get(normalized)
            if canonical and canonical not in used:
                mapping[header] = canonical
                used.add(canonical)
        return mapping

    @staticmethod
    def _normalize_units(dataset: ElectricalDataset, headers: list[str], mapping: dict[str, str]) -> None:
        raw_for = {canonical: raw for raw, canonical in mapping.items()}
        if dataset.has("current_ma"):
            raw = normalize_header(raw_for.get("current_ma", ""))
            values = dataset.numeric("current_ma")
            if raw.endswith("_a") and not raw.endswith("_ma"):
                dataset.columns["current_ma"] = values * 1000.0
        if dataset.has("voltage_mv"):
            raw = normalize_header(raw_for.get("voltage_mv", ""))
            values = dataset.numeric("voltage_mv")
            if raw.endswith("_v") and not raw.endswith("_mv"):
                dataset.columns["voltage_mv"] = values * 1000.0
        if dataset.has("phase_deg") and not dataset.has("phase_mrad"):
            dataset.columns["phase_mrad"] = np.deg2rad(dataset.numeric("phase_deg")) * 1000.0

    @staticmethod
    def _validate_path(path: Path) -> None:
        if not path.is_file():
            raise FileNotFoundError(path)
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported electrical file extension {path.suffix!r}. Supported exports: "
                + ", ".join(sorted(SUPPORTED_EXTENSIONS))
            )


def _read_text_prefix(path: Path, limit: int) -> str:
    with path.open("rb") as stream:
        raw = stream.read(limit)
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _detect_delimiter(sample: str) -> str | None:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        first = next((line for line in sample.splitlines() if line.strip()), "")
        counts = {delimiter: first.count(delimiter) for delimiter in (",", ";", "\t", "|")}
        delimiter, count = max(counts.items(), key=lambda item: item[1])
        return delimiter if count else None


def _to_float(value: Any) -> float:
    if value is None:
        return float("nan")
    if isinstance(value, (int, float, np.number)):
        return float(value)
    text = str(value).strip().replace("−", "-")
    if not text:
        return float("nan")
    # Decimal comma is supported when there is no decimal point.
    if text.count(",") == 1 and "." not in text:
        text = text.replace(",", ".")
    text = re.sub(r"[^0-9eE+\-.]", "", text)
    try:
        return float(text)
    except ValueError:
        return float("nan")


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "base", "reference", "ref"}
