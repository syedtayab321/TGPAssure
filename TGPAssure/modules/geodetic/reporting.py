from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from modules.geodetic.models import FIELD_LABELS, GeodeticDataset, GeodeticQcResult


def _selected_rows(dataset: GeodeticDataset, selected_fields: set[str] | None = None):
    for record in dataset.records:
        values = record.values
        keys = [key for key in values if selected_fields is None or key in selected_fields]
        if keys:
            yield record, keys


def export_selected_text(dataset: GeodeticDataset, path: str | Path, selected_fields: set[str] | None = None) -> Path:
    target = Path(path)
    lines = [f"TGPAssure Geodetic/DC Examiner Export", f"Source: {dataset.source_path}", ""]
    for record, keys in _selected_rows(dataset, selected_fields):
        lines.append(f"Record {record.record_id} | line {record.line_number}")
        for key in keys:
            lines.append(f"  {FIELD_LABELS.get(key, key)}: {record.values.get(key, '')}")
        lines.append("")
    target.write_text("\n".join(lines), encoding="utf-8")
    return target


def export_selected_xlsx(dataset: GeodeticDataset, path: str | Path, selected_fields: set[str] | None = None) -> Path:
    target = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "DC Examiner"
    header = ["Record ID", "Source Line", "Field", "Value"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173A52")
    for record, keys in _selected_rows(dataset, selected_fields):
        for key in keys:
            ws.append([record.record_id, record.line_number, FIELD_LABELS.get(key, key), record.values.get(key, "")])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 28
    wb.save(target)
    return target


def export_qc_pdf(dataset: GeodeticDataset, qc: GeodeticQcResult, path: str | Path) -> Path:
    target = Path(path)
    doc = SimpleDocTemplate(str(target), pagesize=landscape(A4), rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("TGPAssure Geodetic / GNSS Quality-Control Report", styles["Title"]),
        Paragraph(f"Source: {dataset.source_path.name}", styles["Normal"]),
        Paragraph(f"QC profile: {qc.profile_name} | Status: <b>{qc.status}</b> | Score: {qc.score:.1f}/100", styles["Normal"]),
        Spacer(1, 10),
    ]
    metric_rows = [["Metric", "Count", "Minimum", "Median", "Maximum", "Criterion"]]
    for metric in qc.metrics.values():
        finite = metric.finite
        if finite.size:
            criterion = ""
            if metric.threshold is not None:
                criterion = (">= " if metric.direction == "min" else "<= ") + f"{metric.threshold:g} {metric.unit}".strip()
            metric_rows.append([
                metric.label, str(finite.size), f"{finite.min():.6g}", f"{float(__import__('numpy').median(finite)):.6g}",
                f"{finite.max():.6g}", criterion,
            ])
    table = Table(metric_rows, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173A52")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C8D4DE")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
    ]))
    story.extend([Paragraph("QC Metrics", styles["Heading2"]), table, Spacer(1, 12)])

    finding_rows = [["Severity", "Finding", "Evidence / Message", "Action"]]
    for finding in qc.findings:
        finding_rows.append([finding.severity, finding.title, finding.message, finding.suggested_action])
    if len(finding_rows) == 1:
        finding_rows.append(["PASS", "No findings", "No configured QC finding was raised.", ""])
    findings_table = Table(finding_rows, colWidths=[45, 130, 330, 240], repeatRows=1, hAlign="LEFT")
    findings_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173A52")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C8D4DE")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
    ]))
    story.extend([
        Paragraph("Findings", styles["Heading2"]), findings_table, Spacer(1, 10),
        Paragraph(
            "Thresholds in this report are configurable screening criteria. The governing client specification, survey procedure, datum/control requirements and competent surveyor review remain authoritative.",
            styles["Italic"],
        ),
    ])
    doc.build(story)
    return target
