from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np

from modules.gravity.constants import RAW_GRAVITY, TIDE_CORRECTION, TERRAIN_CORRECTION, SUPPORTED_EXTENSIONS
from modules.gravity.models import GravityDataRole, GravityDataset, GravitySurveyType


_ALIASES: dict[str, tuple[str, ...]] = {
    "timestamp": ("timestamp", "datetime", "date_time", "time", "utc", "reading_time"),
    "date": ("date", "reading_date"),
    "station_id": ("station", "station_id", "stationno", "station_no", "point", "point_id", "id"),
    "line_id": ("line", "line_id", "lineno", "line_no", "profile", "profile_id"),
    "latitude": ("latitude", "lat", "y_lat", "lat_deg", "latitude_deg"),
    "longitude": ("longitude", "lon", "long", "lng", "x_lon", "long_deg", "longitude_deg"),
    "x": ("x", "easting", "east", "utm_e", "x_coord", "xcoordinate", "x_coordinate", "east_m"),
    "y": ("y", "northing", "north", "utm_n", "y_coord", "ycoordinate", "y_coordinate", "north_m"),
    "elevation": ("elevation_m", "elevation", "elev", "height", "rl", "z", "altitude", "elev_m", "height_m"),
    RAW_GRAVITY: (
        "observed_gravity_mgal", "gravity", "observed_gravity", "gobs", "raw_gravity", "reading",
        "gravity_mgal", "g_mgal", "obs_gravity", "absolute_gravity", "meter_reading", "mgal",
    ),
    TIDE_CORRECTION: ("tide_correction_mgal", "tide", "tidal", "earth_tide", "tide_corr"),
    TERRAIN_CORRECTION: ("terrain_correction_mgal", "terrain", "terrain_corr", "tc"),
}


def _clean(name: Any) -> str:
    text = str(name or "").strip().lower()
    for token in ("\ufeff", "(", ")", "[", "]", "{", "}", "."):
        text = text.replace(token, "")
    for token in (" ", "-", "/", "\\", ":"):
        text = text.replace(token, "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def _first(mapping: dict[str, str], canonical: str) -> str | None:
    for alias in _ALIASES.get(canonical, (canonical,)):
        key = _clean(alias)
        if key in mapping:
            return mapping[key]
    return None


def _float(value: Any) -> float:
    if value is None:
        return np.nan
    text = str(value).strip().replace(",", "")
    if not text:
        return np.nan
    try:
        return float(text)
    except (TypeError, ValueError):
        return np.nan


def _timestamp(value: Any, date_value: Any = None, index: int = 0) -> np.datetime64:
    candidates = []
    if date_value not in (None, "") and value not in (None, ""):
        candidates.append(f"{date_value} {value}")
    if value not in (None, ""):
        candidates.append(str(value))
    for text in candidates:
        text = text.strip().replace("Z", "+00:00")
        try:
            return np.datetime64(datetime.fromisoformat(text).replace(tzinfo=None), "ms")
        except Exception:
            pass
        for fmt in (
            "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%H:%M:%S",
        ):
            try:
                parsed = datetime.strptime(text, fmt)
                if fmt == "%H:%M:%S":
                    parsed = datetime(1970, 1, 1, parsed.hour, parsed.minute, parsed.second)
                return np.datetime64(parsed, "ms")
            except ValueError:
                continue
    return np.datetime64("1970-01-01T00:00:00", "ms") + np.timedelta64(index, "s")


class GravityReader:
    def inspect(self, path: str | Path) -> dict[str, Any]:
        source = Path(path)
        rows = self._read_rows(source, limit=25)
        if not rows:
            raise ValueError("Gravity input contains no data rows")
        normalized = {_clean(key): key for key in rows[0]}
        mapped = {key: _first(normalized, key) for key in _ALIASES}
        recognized = {k: v for k, v in mapped.items() if v}
        return {
            "source_path": str(source),
            "extension": source.suffix.lower(),
            "columns": list(rows[0].keys()),
            "mapped_fields": recognized,
            "is_gravity_candidate": RAW_GRAVITY in recognized,
            "has_coordinates": bool(recognized.get("latitude") and recognized.get("longitude") or recognized.get("x") and recognized.get("y")),
            "has_elevation": bool(recognized.get("elevation")),
            "preview": rows[:10],
        }

    def read(self, path: str | Path) -> list[dict[str, Any]]:
        """Backward-compatible normalized row reader."""
        dataset = self.read_observations(path)
        rows: list[dict[str, Any]] = []
        for i in range(dataset.record_count):
            row = {
                "timestamp": str(dataset.timestamps[i]),
                "station_id": str(dataset.station_id[i]),
                "line_id": str(dataset.line_id[i]),
                "latitude": float(dataset.latitude[i]),
                "longitude": float(dataset.longitude[i]),
                "elevation_m": float(dataset.elevation[i]),
            }
            row.update({name: float(values[i]) for name, values in dataset.channels.items()})
            rows.append(row)
        return rows

    def read_observations(
        self,
        path: str | Path,
        survey_type: str | GravitySurveyType = GravitySurveyType.LAND,
        crs: str | None = None,
    ) -> GravityDataset:
        return self._read_dataset(path, GravityDataRole.OBSERVATIONS, GravitySurveyType(survey_type), crs=crs)

    def read_base(self, path: str | Path, crs: str | None = None) -> GravityDataset:
        return self._read_dataset(path, GravityDataRole.BASE, GravitySurveyType.BASE_STATION, crs=crs)

    def _read_dataset(self, path: str | Path, role: GravityDataRole, survey_type: GravitySurveyType, crs: str | None) -> GravityDataset:
        source = Path(path).expanduser().resolve()
        if source.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported gravity file type: {source.suffix or '(none)'}")
        rows = self._read_rows(source)
        if not rows:
            raise ValueError("Gravity input contains no data rows")
        mapping = {_clean(key): key for key in rows[0]}
        columns = {key: _first(mapping, key) for key in _ALIASES}
        gravity_col = columns.get(RAW_GRAVITY)
        if not gravity_col:
            raise ValueError("Gravity input requires an observed/raw gravity column (for example gravity, gobs, observed_gravity_mgal)")
        if role == GravityDataRole.OBSERVATIONS and not columns.get("elevation"):
            # Allow imported databases without elevation to open for inspection/map display.
            # Reduction/QC stages will flag the missing elevation instead of leaving the UI stuck at loading.
            columns["elevation"] = None
        if role == GravityDataRole.OBSERVATIONS and not ((columns.get("latitude") and columns.get("longitude")) or (columns.get("x") and columns.get("y"))):
            raise ValueError("Gravity observations require coordinate pairs: latitude/longitude or x/y/easting/northing")

        n = len(rows)
        timestamps = np.empty(n, dtype="datetime64[ms]")
        station = np.empty(n, dtype=object)
        line = np.empty(n, dtype=object)
        latitude = np.full(n, np.nan)
        longitude = np.full(n, np.nan)
        x = np.full(n, np.nan)
        y = np.full(n, np.nan)
        elevation = np.full(n, np.nan)
        raw = np.full(n, np.nan)
        tide = np.zeros(n, dtype=float)
        terrain = np.zeros(n, dtype=float)
        time_col = columns.get("timestamp")
        date_col = columns.get("date")
        for i, row in enumerate(rows):
            timestamps[i] = _timestamp(row.get(time_col) if time_col else None, row.get(date_col) if date_col else None, i)
            station_col = columns.get("station_id")
            line_col = columns.get("line_id")
            station_value = row.get(station_col, "") if station_col else ""
            line_value = row.get(line_col, "") if line_col else ""
            station[i] = str(station_value or f"S{i + 1}").strip() or f"S{i + 1}"
            line[i] = str(line_value or "").strip()
            latitude[i] = _float(row.get(columns.get("latitude"))) if columns.get("latitude") else np.nan
            longitude[i] = _float(row.get(columns.get("longitude"))) if columns.get("longitude") else np.nan
            x[i] = _float(row.get(columns.get("x"))) if columns.get("x") else np.nan
            y[i] = _float(row.get(columns.get("y"))) if columns.get("y") else np.nan
            elevation[i] = _float(row.get(columns.get("elevation"))) if columns.get("elevation") else np.nan
            raw[i] = _float(row.get(gravity_col))
            if columns.get(TIDE_CORRECTION):
                tide[i] = _float(row.get(columns[TIDE_CORRECTION]))
            if columns.get(TERRAIN_CORRECTION):
                terrain[i] = _float(row.get(columns[TERRAIN_CORRECTION]))
        if not np.any(np.isfinite(raw)):
            raise ValueError("Observed gravity column contains no numeric values")
        channels = {RAW_GRAVITY: raw}
        if columns.get(TIDE_CORRECTION):
            channels[TIDE_CORRECTION] = np.nan_to_num(tide, nan=0.0)
        if columns.get(TERRAIN_CORRECTION):
            channels[TERRAIN_CORRECTION] = np.nan_to_num(terrain, nan=0.0)
        inferred_crs = crs
        if inferred_crs is None:
            inferred_crs = "EPSG:4326" if np.any(np.isfinite(latitude) & np.isfinite(longitude)) else "LOCAL/PROJECTED"
        return GravityDataset(
            source,
            role,
            survey_type,
            timestamps,
            channels,
            latitude=latitude,
            longitude=longitude,
            x=x,
            y=y,
            elevation=elevation,
            station_id=station,
            line_id=line,
            is_base=np.full(n, role == GravityDataRole.BASE, dtype=bool),
            metadata={
                "column_mapping": {k: v for k, v in columns.items() if v},
                "imported_at": datetime.now(timezone.utc).isoformat(),
                "missing_elevation": not bool(columns.get("elevation")),
            },
            crs=inferred_crs,
        )

    def _read_rows(self, path: Path, limit: int | None = None) -> list[dict[str, Any]]:
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("openpyxl is required to read Excel gravity files") from exc
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(iterator)]
            except StopIteration:
                return []
            rows = []
            for values in iterator:
                row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
                if any(value not in (None, "") for value in row.values()):
                    rows.append(row)
                if limit and len(rows) >= limit:
                    break
            workbook.close()
            return rows
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        sample = text[:16384]
        lines = [line.strip() for line in text.splitlines() if line.strip() and not line.lstrip().startswith(("#", "//"))]
        if not lines:
            return []
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = self._guess_delimiter(lines[:20])
        if delimiter is None:
            headers = self._split_whitespace_header(lines[0])
            rows: list[dict[str, Any]] = []
            for line in lines[1:]:
                values = line.split()
                if not values:
                    continue
                row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
                if any(value not in (None, "") for value in row.values()):
                    rows.append(row)
                if limit and len(rows) >= limit:
                    break
            return rows
        with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as stream:
            reader = csv.DictReader((line for line in stream if line.strip() and not line.lstrip().startswith(("#", "//"))), delimiter=delimiter)
            rows = []
            for row in reader:
                clean_row = {str(key or "").strip(): value for key, value in dict(row).items() if key is not None}
                if clean_row and any(value not in (None, "") for value in clean_row.values()):
                    rows.append(clean_row)
                if limit and len(rows) >= limit:
                    break
            return rows

    @staticmethod
    def _guess_delimiter(lines: list[str]) -> str | None:
        scores = []
        for delimiter in (",", ";", "\t", "|"):
            counts = [len(line.split(delimiter)) for line in lines if delimiter in line]
            if counts:
                scores.append((max(counts), -abs(max(counts) - min(counts)), delimiter))
        if scores:
            best = max(scores)
            if best[0] > 1:
                return best[2]
        return None

    @staticmethod
    def _split_whitespace_header(line: str) -> list[str]:
        headers = [value.strip() for value in line.split() if value.strip()]
        if not headers:
            return ["col_1"]
        return headers
