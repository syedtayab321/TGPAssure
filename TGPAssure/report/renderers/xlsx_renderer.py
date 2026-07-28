from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Set

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report.report_model import ChartSection, HeadingSection, ReportModel, ReportSectionType, TableSection, TextSection


class XlsxRenderer:
    def __init__(self) -> None:
        self.header_fill = PatternFill("solid", fgColor="173A5E")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_fill = PatternFill("solid", fgColor="DCE6F1")
        self.thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

    def render(self, model: ReportModel, output_path: Path) -> Path:
        output_path = Path(output_path).expanduser().resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        used_names: Set[str] = set()

        summary = workbook.create_sheet("Summary")
        used_names.add("Summary")
        summary["A1"] = model.title
        summary["A1"].font = Font(size=18, bold=True, color="173A5E")
        row = 3
        for key, value in model.metadata.items():
            summary.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
            summary.cell(row=row, column=2, value=self._cell_value(value))
            row += 1
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 90

        sections = sorted(model.sections, key=lambda item: item.order)
        if sections:
            row += 1
            summary.cell(row=row, column=1, value="Report Contents").font = Font(size=13, bold=True, color="173A5E")
            row += 1
            for column, label in enumerate(("Section", "Type / Location"), 1):
                cell = summary.cell(row=row, column=column, value=label)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.thin_border
            contents_row = row + 1
        else:
            contents_row = row

        detail_row = max(contents_row + len(sections) + 1, row + 2)
        for offset, section in enumerate(sections):
            index_row = contents_row + offset
            summary.cell(row=index_row, column=1, value=section.title).border = self.thin_border
            location = "Summary"
            target_sheet = None
            if section.section_type == ReportSectionType.TABLE:
                target_sheet = workbook.create_sheet(self._unique_sheet_name(section.title, used_names))
                self._write_table(target_sheet, section)
                location = target_sheet.title
            elif section.section_type == ReportSectionType.CHART:
                target_sheet = workbook.create_sheet(self._unique_sheet_name(section.title, used_names))
                self._write_chart(target_sheet, section)
                location = target_sheet.title
            elif section.section_type == ReportSectionType.HEADING:
                summary.cell(row=detail_row, column=1, value=section.title).font = Font(
                    size=15 if getattr(section, "level", 1) <= 1 else 12, bold=True, color="173A5E"
                )
                summary.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=2)
                detail_row += 2
            elif section.section_type == ReportSectionType.TEXT:
                summary.cell(row=detail_row, column=1, value=section.title).font = Font(bold=True, color="173A5E")
                detail_row += 1
                summary.cell(row=detail_row, column=1, value=section.content)
                summary.cell(row=detail_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
                summary.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=2)
                detail_row += 2
            type_cell = summary.cell(row=index_row, column=2, value=location)
            type_cell.border = self.thin_border
            if target_sheet is not None:
                type_cell.hyperlink = f"#'{target_sheet.title}'!A1"
                type_cell.style = "Hyperlink"

        workbook.save(output_path)
        return output_path

    @staticmethod
    def _unique_sheet_name(title: str, used: Set[str]) -> str:
        base = "".join(char for char in (title or "Sheet") if char not in "[]:*?/\\")[:31] or "Sheet"
        name = base
        counter = 2
        while name in used:
            suffix = f"_{counter}"
            name = f"{base[:31-len(suffix)]}{suffix}"
            counter += 1
        used.add(name)
        return name

    def _write_table(self, sheet: Any, section: TableSection) -> None:
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = section.title
        sheet["A1"].font = Font(size=15, bold=True, color="173A5E")
        row = 3
        if section.headers:
            for column, header in enumerate(section.headers, 1):
                cell = sheet.cell(row=row, column=column, value=self._cell_value(header))
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = self.thin_border
            row += 1
        for data_row in section.rows:
            for column, value in enumerate(data_row, 1):
                cell = sheet.cell(row=row, column=column, value=self._cell_value(value))
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        sheet.freeze_panes = "A4" if section.headers else "A3"
        max_columns = max(len(section.headers), max((len(item) for item in section.rows), default=1))
        for column in range(1, max_columns + 1):
            values = [str(sheet.cell(row=r, column=column).value or "") for r in range(1, min(sheet.max_row, 200) + 1)]
            width = min(70, max(12, max((len(value) for value in values), default=12) + 2))
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.auto_filter.ref = sheet.dimensions

    def _write_chart(self, sheet: Any, section: ChartSection) -> None:
        labels = [str(value) for value in section.data.get("labels", [])]
        values = [float(value) for value in section.data.get("values", [])]
        sheet["A1"] = section.title
        sheet["A1"].font = Font(size=15, bold=True, color="173A5E")
        sheet["A3"] = "Category"
        sheet["B3"] = "Value"
        for cell in (sheet["A3"], sheet["B3"]):
            cell.fill = self.header_fill
            cell.font = self.header_font
        for row, (label, value) in enumerate(zip(labels, values), 4):
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=value)
        if values:
            chart_type = str(section.chart_type or "bar").lower()
            data_ref = Reference(sheet, min_col=2, min_row=3, max_row=3 + len(values))
            category_ref = Reference(sheet, min_col=1, min_row=4, max_row=3 + len(values))
            if chart_type in {"pie", "donut"}:
                chart = PieChart()
                chart.title = section.title
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(category_ref)
                chart.dLbls = DataLabelList()
                chart.dLbls.showPercent = True
                chart.dLbls.showVal = True
                chart.height = 10
                chart.width = 15
            else:
                chart = BarChart()
                chart.type = "bar" if chart_type in {"horizontal_bar", "barh", "hbar"} or len(labels) > 8 else "col"
                chart.style = 10
                chart.title = section.title
                if chart.type == "bar":
                    chart.x_axis.title = section.x_label or "Value"
                    chart.y_axis.title = section.y_label or "Category"
                else:
                    chart.x_axis.title = section.x_label or "Category"
                    chart.y_axis.title = section.y_label or "Value"
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(category_ref)
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = True
                chart.height = max(9, min(16, 6 + len(values) * 0.45))
                chart.width = 20
            sheet.add_chart(chart, "D3")
        sheet.column_dimensions["A"].width = 45
        sheet.column_dimensions["B"].width = 15

    @staticmethod
    def _cell_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        return str(value)