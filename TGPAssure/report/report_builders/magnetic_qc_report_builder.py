from __future__ import annotations

import io
import json
import math
from pathlib import Path
from typing import Any


class MagneticQcReportBuilder:
    """Build detailed magnetic QC reports with tables, plots and audit detail."""

    STATUS_ORDER = ("PASS", "WARN", "FAIL", "SKIPPED", "ERROR", "PENDING")

    def build(self, result: dict[str, Any], output_path: str | Path, fmt: str) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        normalized = fmt.strip().lower()
        if normalized == "pdf":
            self._build_pdf(result, output)
        elif normalized in {"xlsx", "excel"}:
            self._build_xlsx(result, output)
        else:
            raise ValueError(f"Unsupported magnetic report format: {fmt}")
        return output

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------
    def _build_pdf(self, result: dict[str, Any], output: Path) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Image,
            KeepTogether,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="MagBody", parent=styles["BodyText"], fontSize=8.5, leading=12, spaceAfter=4))
        styles.add(ParagraphStyle(name="MagSmall", parent=styles["BodyText"], fontSize=7.4, leading=9.2))
        styles.add(ParagraphStyle(name="MagTiny", parent=styles["BodyText"], fontSize=6.7, leading=8))
        styles.add(ParagraphStyle(name="MagSection", parent=styles["Heading1"], fontSize=15, leading=18, spaceBefore=6, spaceAfter=6))
        styles.add(ParagraphStyle(name="MagSub", parent=styles["Heading2"], fontSize=11, leading=14, spaceBefore=4, spaceAfter=4))
        styles.add(ParagraphStyle(name="MagNote", parent=styles["BodyText"], fontSize=7.4, leading=9, textColor=colors.HexColor("#4A5568")))

        summary = result.get("summary", {}) or {}
        rover = summary.get("rover", {}) or {}
        base = summary.get("base", {}) or {}
        boundary = summary.get("boundary", {}) or {}
        stages = list(result.get("stage_outcomes", []) or [])
        findings = self._collect_findings(stages)
        status = str(result.get("status", "UNKNOWN")).upper()
        score = self._safe_float(result.get("score"), 0.0)
        status_counts = self._status_counts(stages)
        severity_counts = self._severity_counts(findings)

        document = SimpleDocTemplate(
            str(output),
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title="TGPAssure Magnetic QC Report",
            author="TGPAssure",
            subject="Magnetic Survey Quality Control",
        )

        story: list[Any] = [
            Paragraph("TGPAssure Magnetic Quality-Control Report", styles["Title"]),
            Paragraph(
                "Integrated review of magnetic acquisition integrity, GPS/navigation, sensor performance, base-station stability, diurnal correction, repeatability, line/tie consistency, leveling, processing provenance and grid readiness.",
                styles["MagBody"],
            ),
            Spacer(1, 3 * mm),
        ]

        overview_rows = [
            ["Run UUID", result.get("run_uuid", "")],
            ["QC profile", result.get("profile_name", "")],
            ["Overall status", status],
            ["Overall score", f"{score:.1f}/100"],
            ["Source", rover.get("source_path", "")],
            ["Survey / acquisition type", rover.get("survey_type", rover.get("acquisition_type", ""))],
            ["Records", f"{self._safe_int(rover.get('record_count')):,}"],
            ["Lines", self._safe_int(rover.get("line_count"))],
            ["CRS", rover.get("crs", "") or "Not defined"],
            ["Working CRS", rover.get("working_crs", rover.get("metric_crs", "")) or "Not defined"],
            ["Magnetic units", rover.get("magnetic_units", rover.get("units", "nT")) or "nT"],
            ["Channels", ", ".join(rover.get("channels", []) or []) or "Not reported"],
            ["Acquisition period", f"{rover.get('start_time', '')}  →  {rover.get('end_time', '')}"],
        ]
        story.append(self._key_value_table(overview_rows, styles, 52 * mm, 210 * mm))
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("Executive QC Assessment", styles["MagSection"]))
        story.append(Paragraph(self._executive_summary_text(status, score, status_counts, severity_counts), styles["MagBody"]))
        story.append(self._acceptance_table(status, score, status_counts, severity_counts, styles))

        charts = [
            self._plot_status_counts(status_counts, Image),
            self._plot_severity_counts(severity_counts, Image),
            self._plot_stage_metric_score(stages, Image),
            self._plot_stage_durations(stages, Image),
        ]
        charts = [c for c in charts if c is not None]
        if charts:
            story.extend([Spacer(1, 4 * mm), Paragraph("QC Overview Charts", styles["MagSection"])])
            for i in range(0, len(charts), 2):
                pair = charts[i:i + 2]
                row = pair + ([""] if len(pair) == 1 else [])
                table = Table([row], colWidths=[132 * mm, 132 * mm])
                table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
                story.append(table)
                story.append(Spacer(1, 3 * mm))

        story.append(PageBreak())
        story.append(Paragraph("Dataset Inventory and Acquisition Context", styles["MagSection"]))
        inventory_rows = [
            ["Source file", rover.get("source_path", "")],
            ["Format", rover.get("format", rover.get("source_format", ""))],
            ["Survey / acquisition type", rover.get("survey_type", rover.get("acquisition_type", ""))],
            ["Data role", rover.get("data_role", rover.get("role", ""))],
            ["Sensor / instrument", rover.get("instrument", rover.get("sensor_model", rover.get("instrument_model", "")))],
            ["Sensor serial", rover.get("sensor_serial", rover.get("instrument_serial", ""))],
            ["Logger serial", rover.get("logger_serial", "")],
            ["Records / lines", f"{self._safe_int(rover.get('record_count')):,} / {self._safe_int(rover.get('line_count')):,}"],
            ["Source CRS", rover.get("crs", "") or "Not defined"],
            ["Recommended working CRS", rover.get("working_crs", rover.get("metric_crs", "")) or "Not defined"],
            ["Coordinate validity", self._format_value(rover.get("coordinate_valid_pct", rover.get("valid_coordinate_pct", "Not reported")))],
            ["Channels", ", ".join(rover.get("channels", []) or []) or "Not reported"],
            ["Base station", base.get("source_path", "Not supplied") if base else "Not supplied"],
            ["Boundary", boundary.get("source_path", boundary.get("name", "Not supplied")) if boundary else "Not supplied"],
        ]
        story.append(self._key_value_table(inventory_rows, styles, 66 * mm, 196 * mm))

        channel_stats = summary.get("channel_statistics") or rover.get("channel_statistics") or {}
        channel_rows = self._records_to_rows(channel_stats)
        if channel_rows:
            story.extend([Spacer(1, 4 * mm), Paragraph("Channel Statistics", styles["MagSub"]), self._dynamic_records_table(channel_rows, styles)])

        capability_rows = self._qc_capability_rows(summary, rover, base, boundary, stages)
        story.extend([
            Spacer(1, 4 * mm),
            Paragraph("QC Capability / Data Availability", styles["MagSub"]),
            self._standard_table([["QC area", "Availability", "Reason / evidence"]] + capability_rows, [66 * mm, 38 * mm, 158 * mm], styles, repeat_rows=1),
        ])

        story.append(PageBreak())
        story.append(Paragraph("QC Stage Results", styles["MagSection"]))
        stage_rows = [["#", "Stage", "Status", "Duration", "Findings", "Summary"]]
        for order, stage in enumerate(stages, start=1):
            stage_rows.append([
                order,
                stage.get("display_name") or stage.get("stage_key") or "Unknown",
                str(stage.get("status", "")).upper(),
                f"{self._safe_int(stage.get('duration_ms')):,} ms",
                len(stage.get("findings", []) or []),
                Paragraph(str(stage.get("message", "") or ""), styles["MagSmall"]),
            ])
        story.append(self._standard_table(stage_rows, [11 * mm, 52 * mm, 22 * mm, 24 * mm, 18 * mm, 135 * mm], styles, repeat_rows=1))

        domain_sections = [
            ("Line Geometry and Line Statistics", summary.get("line_statistics") or {}),
            ("Base Station Stability", summary.get("base_statistics") or {}),
            ("Diurnal Correction Statistics", summary.get("diurnal_statistics") or {}),
            ("Repeat Station Statistics", summary.get("repeat_statistics") or {}),
            ("Tie-Line / Intersection Statistics", summary.get("tie_statistics") or summary.get("tie_line_statistics") or {}),
            ("Leveling Statistics", summary.get("leveling_statistics") or {}),
            ("Grid / Product Statistics", summary.get("grid_statistics") or {}),
            ("Processing History", summary.get("processing_history") or summary.get("correction_history") or []),
        ]
        for title, data in domain_sections:
            rows = self._records_to_rows(data)
            if not rows:
                continue
            story.append(PageBreak())
            story.append(Paragraph(title, styles["MagSection"]))
            story.append(self._dynamic_records_table(rows, styles))
            chart = self._plot_record_values(title, data, Image)
            if chart is not None:
                story.extend([Spacer(1, 4 * mm), chart])

        story.append(PageBreak())
        story.append(Paragraph("Detailed Stage Metrics", styles["MagSection"]))
        for stage in stages:
            metrics = stage.get("metrics", {}) or {}
            if not isinstance(metrics, dict) or not metrics:
                continue
            title = stage.get("display_name") or stage.get("stage_key") or "Stage"
            rows = [["Metric", "Value"]]
            for key, value in sorted(metrics.items(), key=lambda item: str(item[0])):
                rows.append([self._humanize(key), self._format_value(value)])
            story.append(KeepTogether([
                Paragraph(str(title), styles["MagSub"]),
                self._standard_table(rows, [80 * mm, 182 * mm], styles, repeat_rows=1),
                Spacer(1, 3 * mm),
            ]))

        story.append(PageBreak())
        story.append(Paragraph("QC Findings and Corrective Actions", styles["MagSection"]))
        finding_rows = [["Severity", "Stage", "Rule", "Finding", "Location", "Recommended action"]]
        for stage_name, item in findings:
            finding_rows.append([
                str(item.get("severity", "INFO")).upper(),
                stage_name,
                item.get("rule_id", ""),
                Paragraph(str(item.get("message", "") or ""), styles["MagTiny"]),
                Paragraph(str(item.get("location_ref", "") or ""), styles["MagTiny"]),
                Paragraph(str(item.get("suggested_action", "") or ""), styles["MagTiny"]),
            ])
        if len(finding_rows) == 1:
            finding_rows.append(["INFO", "Summary", "—", "No QC findings were generated.", "—", "No corrective action required."])
        story.append(self._standard_table(finding_rows, [20 * mm, 38 * mm, 34 * mm, 74 * mm, 36 * mm, 60 * mm], styles, repeat_rows=1, font_size=6.7))

        story.extend([Spacer(1, 5 * mm), Paragraph("Recommended Actions Before Acceptance", styles["MagSub"])])
        for index, action in enumerate(self._recommendations(findings, status), start=1):
            story.append(Paragraph(f"{index}. {action}", styles["MagBody"]))

        thresholds = self._extract_thresholds(result)
        if thresholds:
            story.append(PageBreak())
            story.append(Paragraph("QC Profile Thresholds", styles["MagSection"]))
            rows = [["Threshold", "Configured value"]] + [[self._humanize(k), self._format_value(v)] for k, v in sorted(thresholds.items())]
            story.append(self._standard_table(rows, [120 * mm, 142 * mm], styles, repeat_rows=1))

        story.extend([
            Spacer(1, 5 * mm),
            Paragraph(
                "Interpretation note: automated magnetic QC identifies acquisition, navigation, sensor, base-station, processing and spatial-consistency issues. A flagged magnetic anomaly may be geological, cultural or acquisition-related; final acceptance and interpretation require competent geophysical review and supporting field documentation.",
                styles["MagNote"],
            ),
        ])

        document.build(story, onFirstPage=self._page_decorator, onLaterPages=self._page_decorator)

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def _build_xlsx(self, result: dict[str, Any], output: Path) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary = result.get("summary", {}) or {}
        rover = summary.get("rover", {}) or {}
        base = summary.get("base", {}) or {}
        boundary = summary.get("boundary", {}) or {}
        stages = list(result.get("stage_outcomes", []) or [])
        findings = self._collect_findings(stages)
        status_counts = self._status_counts(stages)
        severity_counts = self._severity_counts(findings)

        summary_sheet.append(["TGPAssure Magnetic QC Report", ""])
        summary_sheet.append(["Metric", "Value"])
        rows = [
            ("Run UUID", result.get("run_uuid")),
            ("Profile", result.get("profile_name")),
            ("Status", str(result.get("status", "")).upper()),
            ("Score", result.get("score")),
            ("Started", result.get("started_at")),
            ("Completed", result.get("completed_at")),
            ("Rover Source", rover.get("source_path")),
            ("Survey Type", rover.get("survey_type", rover.get("acquisition_type"))),
            ("Record Count", rover.get("record_count")),
            ("Line Count", rover.get("line_count")),
            ("CRS", rover.get("crs")),
            ("Working CRS", rover.get("working_crs", rover.get("metric_crs"))),
            ("Channels", ", ".join(rover.get("channels", []) or [])),
            ("Base Source", base.get("source_path") if base else "Not supplied"),
            ("Boundary", boundary.get("source_path", boundary.get("name")) if boundary else "Not supplied"),
        ]
        for row in rows:
            summary_sheet.append(list(row))

        summary_sheet.append([])
        summary_sheet.append(["Stage Status", "Count"])
        status_start = summary_sheet.max_row + 1
        for key in self.STATUS_ORDER:
            if status_counts.get(key, 0):
                summary_sheet.append([key, status_counts[key]])
        status_end = summary_sheet.max_row

        summary_sheet.append([])
        summary_sheet.append(["Finding Severity", "Count"])
        severity_start = summary_sheet.max_row + 1
        for key, value in severity_counts.items():
            if value:
                summary_sheet.append([key, value])
        severity_end = summary_sheet.max_row

        stages_sheet = workbook.create_sheet("QC Stages")
        stages_sheet.append(["Order", "Stage Key", "Stage", "Status", "Score", "Duration ms", "Finding Count", "Message", "Metrics"])
        findings_sheet = workbook.create_sheet("Findings")
        findings_sheet.append(["Severity", "Stage", "Rule", "Finding", "Location", "Recommended Action", "Metadata"])
        metrics_sheet = workbook.create_sheet("Detailed Metrics")
        metrics_sheet.append(["Stage", "Stage Key", "Metric", "Value"])

        for order, stage in enumerate(stages, start=1):
            metrics = stage.get("metrics", {}) or {}
            stages_sheet.append([
                order,
                stage.get("stage_key"),
                stage.get("display_name"),
                str(stage.get("status", "")).upper(),
                stage.get("score", metrics.get("score", metrics.get("qc_score")) if isinstance(metrics, dict) else None),
                stage.get("duration_ms"),
                len(stage.get("findings", []) or []),
                stage.get("message"),
                json.dumps(metrics, ensure_ascii=False, default=str),
            ])
            if isinstance(metrics, dict):
                for key, value in sorted(metrics.items()):
                    metrics_sheet.append([stage.get("display_name"), stage.get("stage_key"), self._humanize(key), self._format_value(value)])
            for item in stage.get("findings", []) or []:
                findings_sheet.append([
                    str(item.get("severity", "")).upper(),
                    stage.get("display_name"),
                    item.get("rule_id"),
                    item.get("message"),
                    item.get("location_ref"),
                    item.get("suggested_action"),
                    json.dumps(item.get("metadata", {}), ensure_ascii=False, default=str),
                ])

        domain_sheets = [
            ("Line Statistics", summary.get("line_statistics") or {}),
            ("Base Statistics", summary.get("base_statistics") or {}),
            ("Diurnal Statistics", summary.get("diurnal_statistics") or {}),
            ("Repeat Statistics", summary.get("repeat_statistics") or {}),
            ("Tie Statistics", summary.get("tie_statistics") or summary.get("tie_line_statistics") or {}),
            ("Leveling Statistics", summary.get("leveling_statistics") or {}),
            ("Grid Statistics", summary.get("grid_statistics") or {}),
            ("Channel Statistics", summary.get("channel_statistics") or rover.get("channel_statistics") or {}),
            ("Processing History", summary.get("processing_history") or summary.get("correction_history") or []),
        ]
        for title, data in domain_sheets:
            self._write_data_sheet(workbook.create_sheet(title), data)

        thresholds_sheet = workbook.create_sheet("Thresholds")
        thresholds_sheet.append(["Threshold", "Configured Value"])
        for key, value in sorted(self._extract_thresholds(result).items()):
            thresholds_sheet.append([self._humanize(key), self._format_value(value)])

        charts_sheet = workbook.create_sheet("Charts")
        charts_sheet.append(["Charts generated from QC summary tables"])
        if status_end >= status_start:
            chart = PieChart()
            chart.title = "QC Stage Status Distribution"
            data_ref = Reference(summary_sheet, min_col=2, min_row=status_start - 1, max_row=status_end)
            cats_ref = Reference(summary_sheet, min_col=1, min_row=status_start, max_row=status_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.dLbls = DataLabelList()
            chart.dLbls.showPercent = True
            chart.dLbls.showVal = True
            chart.height = 10
            chart.width = 14
            charts_sheet.add_chart(chart, "A3")
        if severity_end >= severity_start:
            chart2 = PieChart()
            chart2.title = "Findings by Severity"
            data_ref = Reference(summary_sheet, min_col=2, min_row=severity_start - 1, max_row=severity_end)
            cats_ref = Reference(summary_sheet, min_col=1, min_row=severity_start, max_row=severity_end)
            chart2.add_data(data_ref, titles_from_data=True)
            chart2.set_categories(cats_ref)
            chart2.dLbls = DataLabelList()
            chart2.dLbls.showPercent = True
            chart2.dLbls.showVal = True
            chart2.height = 10
            chart2.width = 14
            charts_sheet.add_chart(chart2, "A22")
        if stages_sheet.max_row >= 2:
            score_chart = BarChart()
            score_chart.type = "bar"
            score_chart.title = "QC Stage Scores"
            score_chart.x_axis.title = "Score"
            score_chart.y_axis.title = "Stage"
            score_chart.add_data(Reference(stages_sheet, min_col=5, min_row=1, max_row=stages_sheet.max_row), titles_from_data=True)
            score_chart.set_categories(Reference(stages_sheet, min_col=3, min_row=2, max_row=stages_sheet.max_row))
            score_chart.dLbls = DataLabelList()
            score_chart.dLbls.showVal = True
            score_chart.height = max(10, min(17, 7 + stages_sheet.max_row * 0.35))
            score_chart.width = 22
            charts_sheet.add_chart(score_chart, "P3")

            duration_chart = BarChart()
            duration_chart.type = "bar"
            duration_chart.title = "QC Stage Execution Duration"
            duration_chart.x_axis.title = "Duration (ms)"
            duration_chart.y_axis.title = "Stage"
            duration_chart.add_data(Reference(stages_sheet, min_col=6, min_row=1, max_row=stages_sheet.max_row), titles_from_data=True)
            duration_chart.set_categories(Reference(stages_sheet, min_col=3, min_row=2, max_row=stages_sheet.max_row))
            duration_chart.dLbls = DataLabelList()
            duration_chart.dLbls.showVal = True
            duration_chart.height = max(10, min(17, 7 + stages_sheet.max_row * 0.35))
            duration_chart.width = 22
            charts_sheet.add_chart(duration_chart, "P22")

        for sheet in workbook.worksheets:
            self._format_sheet(sheet)
        summary_sheet.freeze_panes = "A3"
        workbook.save(output)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _page_decorator(canvas, doc) -> None:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.units import mm
        canvas.saveState()
        width, _ = doc.pagesize
        canvas.setStrokeColor(HexColor("#CBD5E0"))
        canvas.line(10 * mm, 9 * mm, width - 10 * mm, 9 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(HexColor("#4A5568"))
        canvas.drawString(10 * mm, 5.5 * mm, "TGPAssure — Magnetic Quality Control")
        canvas.drawRightString(width - 10 * mm, 5.5 * mm, f"Page {doc.page}")
        canvas.restoreState()

    def _key_value_table(self, rows: list[list[Any]], styles, key_width: float, value_width: float):
        from reportlab.lib import colors
        from reportlab.platypus import Paragraph, Table, TableStyle
        data = [[Paragraph(str(k), styles["MagSmall"]), Paragraph(self._format_value(v), styles["MagSmall"])] for k, v in rows]
        table = Table(data, colWidths=[key_width, value_width])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EEF5")),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#AAB5C0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return table

    def _standard_table(self, rows: list[list[Any]], widths: list[float], styles, repeat_rows: int = 0, font_size: float = 7.5):
        from reportlab.lib import colors
        from reportlab.platypus import Paragraph, Table, TableStyle
        converted = []
        for row in rows:
            converted_row = []
            for cell in row:
                if hasattr(cell, "wrap"):
                    converted_row.append(cell)
                else:
                    converted_row.append(Paragraph(self._format_value(cell), styles["MagTiny"] if font_size < 7 else styles["MagSmall"]))
            converted.append(converted_row)
        table = Table(converted, repeatRows=repeat_rows, colWidths=widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#AAB5C0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return table

    def _acceptance_table(self, status: str, score: float, status_counts: dict[str, int], severity_counts: dict[str, int], styles):
        from reportlab.lib.units import mm
        critical = severity_counts.get("CRITICAL", 0) + severity_counts.get("ERROR", 0)
        warnings = severity_counts.get("WARNING", 0) + severity_counts.get("WARN", 0)
        fails = status_counts.get("FAIL", 0) + status_counts.get("ERROR", 0)
        decision = "ACCEPT"
        if status in {"FAIL", "ERROR"} or critical or fails:
            decision = "HOLD / CORRECT BEFORE ACCEPTANCE"
        elif status == "WARN" or warnings:
            decision = "CONDITIONAL ACCEPTANCE — REVIEW WARNINGS"
        rows = [
            ["Decision", "Score", "Pass", "Warn", "Fail/Error", "Critical/Error findings"],
            [decision, f"{score:.1f}/100", status_counts.get("PASS", 0), status_counts.get("WARN", 0), fails, critical],
        ]
        return self._standard_table(rows, [78 * mm, 32 * mm, 30 * mm, 30 * mm, 40 * mm, 52 * mm], styles, repeat_rows=1)

    def _qc_capability_rows(self, summary: dict[str, Any], rover: dict[str, Any], base: dict[str, Any], boundary: dict[str, Any], stages: list[dict[str, Any]]) -> list[list[Any]]:
        stage_text = " ".join(f"{s.get('stage_key', '')} {s.get('display_name', '')}" for s in stages).lower()
        rows = [
            ["File / sensor integrity", "Available", "Magnetic measurements present"],
            ["GPS / coordinate QC", "Available" if rover.get("crs") or rover.get("coordinate_valid_pct") is not None else "Limited", rover.get("crs") or "Coordinate CRS not reported"],
            ["Boundary QC", "Available" if boundary else "Skipped / optional", boundary.get("source_path", "No boundary supplied") if boundary else "No survey boundary supplied"],
            ["Base / diurnal QC", "Available" if base else "Skipped / optional", base.get("source_path", "") if base else "No separate base-station dataset supplied"],
            ["Tie-line / leveling QC", "Available" if ("tie" in stage_text or rover.get("tie_line_count")) else "Conditional", "Requires traverse/tie geometry"],
            ["Grid QC", "Available" if ("grid" in stage_text or summary.get("grid_statistics")) else "Conditional", "Requires sufficient spatial coverage and a valid metric CRS"],
        ]
        return rows

    def _dynamic_records_table(self, rows: list[dict[str, Any]], styles):
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        keys = keys[:8]
        data = [[self._humanize(key) for key in keys]]
        for row in rows[:200]:
            data.append([self._format_value(row.get(key)) for key in keys])
        from reportlab.lib.units import mm
        width = 262.0 / max(len(keys), 1)
        return self._standard_table(data, [width * mm] * len(keys), styles, repeat_rows=1, font_size=6.7)

    def _plot_status_counts(self, counts: dict[str, int], Image):
        labels = [key for key in self.STATUS_ORDER if counts.get(key, 0)]
        values = [counts[key] for key in labels]
        return self._matplotlib_bar("QC Stage Status Distribution", labels, values, "Stages", Image)

    def _plot_severity_counts(self, counts: dict[str, int], Image):
        labels = [key for key, value in counts.items() if value]
        values = [counts[key] for key in labels]
        return self._matplotlib_bar("Findings by Severity", labels, values, "Findings", Image)

    def _plot_stage_metric_score(self, stages: list[dict[str, Any]], Image):
        labels: list[str] = []
        values: list[float] = []
        for stage in stages:
            metrics = stage.get("metrics", {}) or {}
            candidate = stage.get("score")
            if candidate is None and isinstance(metrics, dict):
                candidate = metrics.get("score", metrics.get("qc_score"))
            if candidate is None:
                continue
            value = self._safe_float(candidate, math.nan)
            if math.isfinite(value):
                labels.append(str(stage.get("display_name") or stage.get("stage_key") or "Stage"))
                values.append(value)
        if not values:
            return None
        return self._matplotlib_bar("QC Stage Scores", labels[:20], values[:20], "Score", Image, horizontal=True)

    def _plot_stage_durations(self, stages: list[dict[str, Any]], Image):
        rows = [(str(stage.get("display_name") or stage.get("stage_key") or "Stage"), self._safe_float(stage.get("duration_ms"), 0.0)) for stage in stages]
        rows = [(label, value) for label, value in rows if value > 0]
        if not rows:
            return None
        rows = sorted(rows, key=lambda item: item[1], reverse=True)[:15]
        return self._matplotlib_bar("Longest QC Stages", [x[0] for x in rows], [x[1] for x in rows], "Duration (ms)", Image, horizontal=True)

    def _plot_record_values(self, title: str, data: Any, Image):
        rows = self._records_to_rows(data)
        if len(rows) < 2:
            return None
        numeric_keys: list[str] = []
        for key in rows[0].keys():
            if sum(self._is_number(row.get(key)) for row in rows) >= min(3, len(rows)):
                numeric_keys.append(key)
        if not numeric_keys:
            return None
        value_key = numeric_keys[-1]
        labels = [str(row.get("line_id") or row.get("station_id") or row.get("id") or i + 1) for i, row in enumerate(rows[:30])]
        values = [self._safe_float(row.get(value_key), 0.0) for row in rows[:30]]
        return self._matplotlib_bar(f"{title}: {self._humanize(value_key)}", labels, values, self._humanize(value_key), Image, horizontal=len(labels) > 10)

    @staticmethod
    def _matplotlib_bar(title: str, labels: list[str], values: list[float], ylabel: str, Image, horizontal: bool = False):
        if not labels or not values:
            return None
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except Exception:
            return None
        fig, ax = plt.subplots(figsize=(9.2, 4.2))
        positions = list(range(len(labels)))
        if horizontal:
            ax.barh(positions, values)
            ax.set_yticks(positions)
            ax.set_yticklabels([label[:42] for label in labels], fontsize=7)
            ax.invert_yaxis()
            ax.set_xlabel(ylabel)
            ax.grid(axis="x", alpha=0.25)
        else:
            ax.bar(positions, values)
            ax.set_xticks(positions)
            ax.set_xticklabels([label[:24] for label in labels], rotation=30, ha="right", fontsize=7)
            ax.set_ylabel(ylabel)
            ax.grid(axis="y", alpha=0.25)
        ax.set_title(title)
        fig.tight_layout()
        buffer = io.BytesIO()
        fig.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
        plt.close(fig)
        buffer.seek(0)
        from reportlab.lib.units import mm
        image = Image(buffer, width=126 * mm, height=58 * mm)
        image._tgp_buffer = buffer
        return image

    @staticmethod
    def _collect_findings(stages: list[dict[str, Any]]) -> list[tuple[str, dict[str, Any]]]:
        result: list[tuple[str, dict[str, Any]]] = []
        for stage in stages:
            stage_name = stage.get("display_name") or stage.get("stage_key") or "Unknown"
            result.extend((str(stage_name), item) for item in (stage.get("findings", []) or []))
        return result

    @classmethod
    def _status_counts(cls, stages: list[dict[str, Any]]) -> dict[str, int]:
        counts = {key: 0 for key in cls.STATUS_ORDER}
        for stage in stages:
            key = str(stage.get("status", "UNKNOWN")).upper()
            counts[key] = counts.get(key, 0) + 1
        return counts

    @staticmethod
    def _severity_counts(findings: list[tuple[str, dict[str, Any]]]) -> dict[str, int]:
        counts: dict[str, int] = {}
        for _, finding in findings:
            key = str(finding.get("severity", "INFO")).upper()
            counts[key] = counts.get(key, 0) + 1
        return counts

    def _executive_summary_text(self, status: str, score: float, status_counts: dict[str, int], severity_counts: dict[str, int]) -> str:
        critical = severity_counts.get("CRITICAL", 0) + severity_counts.get("ERROR", 0)
        warnings = severity_counts.get("WARNING", 0) + severity_counts.get("WARN", 0)
        fails = status_counts.get("FAIL", 0) + status_counts.get("ERROR", 0)
        if status in {"FAIL", "ERROR"} or critical or fails:
            decision = "The magnetic dataset should be held for corrective review before final acceptance or interpretation delivery."
        elif status == "WARN" or warnings:
            decision = "The magnetic dataset is conditionally acceptable subject to review of warnings, field notes, base-station support and processing provenance."
        else:
            decision = "No material automated QC failure was detected; the dataset is suitable for technical review toward acceptance."
        return (
            f"Overall automated result: <b>{status}</b> with a score of <b>{score:.1f}/100</b>. "
            f"The pipeline recorded {status_counts.get('PASS', 0)} passing stages, {status_counts.get('WARN', 0)} warning stages and {fails} failed/error stages. "
            f"There are {critical} critical/error findings and {warnings} warning findings. {decision}"
        )

    def _recommendations(self, findings: list[tuple[str, dict[str, Any]]], status: str) -> list[str]:
        actions: list[str] = []
        for _, item in findings:
            action = str(item.get("suggested_action", "") or "").strip()
            if action and action not in actions:
                actions.append(action)
            if len(actions) >= 12:
                break
        if not actions:
            if status in {"FAIL", "ERROR"}:
                actions.append("Review failed stages, source acquisition logs, GPS/sensor status, base-station coverage and processing history before release.")
            elif status == "WARN":
                actions.append("Review warning findings and confirm whether deviations reflect valid geology, cultural interference, GPS limitations or acquisition/processing issues.")
            else:
                actions.append("Retain the QC report with the magnetic deliverables and obtain final geophysicist sign-off.")
        return actions

    def _extract_thresholds(self, result: dict[str, Any]) -> dict[str, Any]:
        candidates = [
            result.get("thresholds"),
            (result.get("parameters") or {}).get("thresholds") if isinstance(result.get("parameters"), dict) else None,
            (result.get("summary") or {}).get("thresholds") if isinstance(result.get("summary"), dict) else None,
            result.get("profile_thresholds"),
        ]
        for candidate in candidates:
            if isinstance(candidate, dict) and candidate:
                return candidate
        return {}

    @staticmethod
    def _records_to_rows(data: Any) -> list[dict[str, Any]]:
        if isinstance(data, list):
            return [row for row in data if isinstance(row, dict)]
        if isinstance(data, dict):
            if data and all(isinstance(value, dict) for value in data.values()):
                rows = []
                for key, value in data.items():
                    row = {"id": key}
                    row.update(value)
                    rows.append(row)
                return rows
            return [{"metric": key, "value": value} for key, value in data.items()]
        return []

    @staticmethod
    def _format_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if math.isnan(value):
                return "NaN"
            if abs(value) >= 1000:
                return f"{value:,.3f}"
            return f"{value:.6g}"
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return str(value)

    @staticmethod
    def _humanize(value: Any) -> str:
        return str(value).replace("_", " ").strip().title()

    @staticmethod
    def _safe_float(value: Any, default: float = 0.0) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _safe_int(value: Any, default: int = 0) -> int:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _is_number(value: Any) -> bool:
        try:
            float(value)
            return value is not None
        except (TypeError, ValueError):
            return False

    def _write_data_sheet(self, sheet, data: Any) -> None:
        rows = self._records_to_rows(data)
        if not rows:
            sheet.append(["No data"])
            return
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        sheet.append([self._humanize(key) for key in keys])
        for row in rows:
            sheet.append([self._format_value(row.get(key)) for key in keys])

    @staticmethod
    def _format_sheet(sheet) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        if sheet.max_row:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2C3E50")
        if sheet.max_row >= 2 and sheet.title != "Summary":
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            width = min(max(max((len(str(cell.value or "")) for cell in sheet[get_column_letter(column)]), default=10) + 2, 12), 60)
            sheet.column_dimensions[get_column_letter(column)].width = width
