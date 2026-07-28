from __future__ import annotations

import io
import json
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


class GravityQcReportBuilder:
    """Build detailed land-gravity QC reports in PDF or Excel format.

    The builder is deliberately defensive: it accepts the existing TGPAssure
    result dictionary and enriches the report whenever optional statistics are
    present, without requiring every QC stage to populate every field.
    """

    STATUS_ORDER = ("PASS", "WARN", "FAIL", "SKIPPED", "ERROR", "PENDING")
    SEVERITY_ORDER = ("CRITICAL", "ERROR", "WARNING", "WARN", "INFO")

    def build(self, result: dict[str, Any], output_path: str | Path, fmt: str) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        normalized = fmt.strip().lower()
        if normalized == "pdf":
            self._build_pdf(result, output)
        elif normalized in {"xlsx", "excel"}:
            self._build_xlsx(result, output)
        else:
            raise ValueError(f"Unsupported gravity report format: {fmt}")
        return output

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------
    def _build_pdf(self, result: dict[str, Any], output: Path) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Image,
            KeepTogether,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="GravBody", parent=styles["BodyText"], fontSize=8.5, leading=12, spaceAfter=4))
        styles.add(ParagraphStyle(name="GravSmall", parent=styles["BodyText"], fontSize=7.4, leading=9.2))
        styles.add(ParagraphStyle(name="GravTiny", parent=styles["BodyText"], fontSize=6.8, leading=8))
        styles.add(ParagraphStyle(name="GravSection", parent=styles["Heading1"], fontSize=15, leading=18, spaceBefore=6, spaceAfter=6))
        styles.add(ParagraphStyle(name="GravSub", parent=styles["Heading2"], fontSize=11, leading=14, spaceBefore=4, spaceAfter=4))
        styles.add(ParagraphStyle(name="GravNote", parent=styles["BodyText"], fontSize=7.4, leading=9, textColor=colors.HexColor("#4A5568")))

        summary = result.get("summary", {}) or {}
        observations = summary.get("observations", {}) or {}
        stages = list(result.get("stage_outcomes", []) or [])
        findings = self._collect_findings(stages)
        status = str(result.get("status", "UNKNOWN")).upper()
        score = self._safe_float(result.get("score"), 0.0)

        document = SimpleDocTemplate(
            str(output),
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title="TGPAssure Gravity QC Report",
            author="TGPAssure",
            subject="Land Gravity Quality Control",
        )

        story: list[Any] = []
        story.extend([
            Paragraph("TGPAssure Land Gravity Quality-Control Report", styles["Title"]),
            Paragraph(
                "Field acquisition, base/drift stability, corrections, repeatability, loop closure, cross-over analysis and final Bouguer-anomaly consistency review.",
                styles["GravBody"],
            ),
            Spacer(1, 3 * mm),
        ])

        status_counts = self._status_counts(stages)
        severity_counts = self._severity_counts(findings)
        overview_rows = [
            ["Run UUID", result.get("run_uuid", "")],
            ["QC profile", result.get("profile_name", "")],
            ["Overall status", status],
            ["Overall score", f"{score:.1f}/100"],
            ["Source", observations.get("source_path", "")],
            ["Survey type", observations.get("survey_type", "")],
            ["Records", f"{self._safe_int(observations.get('record_count')):,}"],
            ["Stations", self._safe_int(observations.get("station_count"))],
            ["Lines", self._safe_int(observations.get("line_count"))],
            ["CRS", observations.get("crs", "") or "Not defined"],
            ["Gravity units", observations.get("gravity_units", "mGal") or "mGal"],
            ["Acquisition period", f"{observations.get('start_time', '')}  →  {observations.get('end_time', '')}"],
            ["Final channels", ", ".join(summary.get("final_channels", []) or []) or "Not reported"],
        ]
        story.append(self._key_value_table(overview_rows, styles, 48 * mm, 214 * mm))
        story.append(Spacer(1, 4 * mm))

        decision_text = self._executive_summary_text(status, score, status_counts, severity_counts, findings)
        story.extend([
            Paragraph("Executive QC Assessment", styles["GravSection"]),
            Paragraph(decision_text, styles["GravBody"]),
            self._acceptance_table(status, score, status_counts, severity_counts, styles),
            Spacer(1, 4 * mm),
        ])

        charts = [
            self._plot_status_counts(status_counts, Image),
            self._plot_severity_counts(severity_counts, Image),
            self._plot_stage_durations(stages, Image),
        ]
        charts = [chart for chart in charts if chart is not None]
        if charts:
            story.append(Paragraph("QC Overview Charts", styles["GravSection"]))
            for i in range(0, len(charts), 2):
                pair = charts[i:i + 2]
                row = pair + ([""] if len(pair) == 1 else [])
                table = Table([row], colWidths=[132 * mm, 132 * mm])
                table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
                story.append(table)
                story.append(Spacer(1, 3 * mm))

        story.append(PageBreak())
        story.append(Paragraph("Survey and Reduction Inventory", styles["GravSection"]))
        inventory = self._inventory_rows(summary, observations)
        story.append(self._key_value_table(inventory, styles, 65 * mm, 197 * mm))

        correction_rows = self._correction_chain_rows(summary, stages)
        if correction_rows:
            story.extend([
                Spacer(1, 4 * mm),
                Paragraph("Gravity Reduction / Correction Chain", styles["GravSub"]),
                self._standard_table(
                    [["Step", "Status / Evidence", "Key metric(s)"]] + correction_rows,
                    [48 * mm, 78 * mm, 136 * mm],
                    styles,
                    repeat_rows=1,
                ),
            ])

        stage_chart = self._plot_stage_metric_score(stages, Image)
        if stage_chart is not None:
            story.extend([Spacer(1, 4 * mm), stage_chart])

        story.append(PageBreak())
        story.append(Paragraph("QC Stage Results", styles["GravSection"]))
        stage_rows = [["#", "Stage", "Status", "Duration", "Findings", "Summary"]]
        for order, stage in enumerate(stages, start=1):
            stage_rows.append([
                order,
                stage.get("display_name") or stage.get("stage_key") or "Unknown",
                str(stage.get("status", "")).upper(),
                f"{self._safe_int(stage.get('duration_ms')):,} ms",
                len(stage.get("findings", []) or []),
                Paragraph(str(stage.get("message", "") or ""), styles["GravSmall"]),
            ])
        story.append(self._standard_table(stage_rows, [11 * mm, 50 * mm, 22 * mm, 25 * mm, 18 * mm, 136 * mm], styles, repeat_rows=1))

        # Domain-specific detailed statistics.
        domain_blocks = [
            ("Base Station & Drift Stability", summary.get("base_statistics") or {}),
            ("Repeat Station Precision", summary.get("repeat_statistics") or {}),
            ("Loop Closure Analysis", summary.get("loop_statistics") or {}),
            ("Cross-over Error Analysis", summary.get("crossover_statistics") or {}),
            ("Terrain / Elevation / Bouguer Statistics", summary.get("reduction_statistics") or summary.get("correction_statistics") or {}),
            ("Final Anomaly Statistics", summary.get("anomaly_statistics") or summary.get("final_anomaly_statistics") or {}),
        ]
        for title, data in domain_blocks:
            rows = self._records_to_rows(data)
            if not rows:
                continue
            story.append(PageBreak())
            story.append(Paragraph(title, styles["GravSection"]))
            story.append(self._dynamic_records_table(rows, styles))
            domain_chart = self._plot_record_values(title, data, Image)
            if domain_chart is not None:
                story.extend([Spacer(1, 4 * mm), domain_chart])

        story.append(PageBreak())
        story.append(Paragraph("Detailed Stage Metrics", styles["GravSection"]))
        for stage in stages:
            metrics = stage.get("metrics", {}) or {}
            if not isinstance(metrics, dict) or not metrics:
                continue
            title = stage.get("display_name") or stage.get("stage_key") or "Stage"
            metric_rows = [["Metric", "Value"]]
            for key, value in sorted(metrics.items(), key=lambda item: str(item[0])):
                metric_rows.append([self._humanize(key), self._format_value(value)])
            block = [
                Paragraph(str(title), styles["GravSub"]),
                self._standard_table(metric_rows, [75 * mm, 187 * mm], styles, repeat_rows=1),
                Spacer(1, 3 * mm),
            ]
            story.append(KeepTogether(block))

        story.append(PageBreak())
        story.append(Paragraph("QC Findings and Corrective Actions", styles["GravSection"]))
        finding_rows = [["Severity", "Stage", "Rule", "Finding", "Location", "Recommended action"]]
        for stage_name, item in findings:
            finding_rows.append([
                str(item.get("severity", "INFO")).upper(),
                stage_name,
                item.get("rule_id", ""),
                Paragraph(str(item.get("message", "") or ""), styles["GravTiny"]),
                Paragraph(str(item.get("location_ref", "") or ""), styles["GravTiny"]),
                Paragraph(str(item.get("suggested_action", "") or ""), styles["GravTiny"]),
            ])
        if len(finding_rows) == 1:
            finding_rows.append(["INFO", "Summary", "—", "No QC findings were generated.", "—", "No corrective action required."])
        story.append(self._standard_table(finding_rows, [20 * mm, 38 * mm, 34 * mm, 74 * mm, 36 * mm, 60 * mm], styles, repeat_rows=1, font_size=6.7))

        recommendations = self._recommendations(findings, status)
        story.extend([
            Spacer(1, 5 * mm),
            Paragraph("Recommended Actions Before Acceptance", styles["GravSub"]),
        ])
        for index, text in enumerate(recommendations, start=1):
            story.append(Paragraph(f"{index}. {text}", styles["GravBody"]))

        thresholds = self._extract_thresholds(result)
        if thresholds:
            story.append(PageBreak())
            story.append(Paragraph("QC Profile Thresholds", styles["GravSection"]))
            threshold_rows = [["Threshold", "Configured value"]] + [
                [self._humanize(key), self._format_value(value)] for key, value in sorted(thresholds.items())
            ]
            story.append(self._standard_table(threshold_rows, [120 * mm, 142 * mm], styles, repeat_rows=1))

        story.extend([
            Spacer(1, 6 * mm),
            Paragraph(
                "Interpretation note: automated QC flags indicate conditions requiring review; they do not by themselves distinguish processing error from a valid geological gravity response. Final acceptance should be performed by a competent geophysicist with field notes, instrument records, elevation-control information and the adopted gravity datum/reduction parameters.",
                styles["GravNote"],
            ),
        ])

        document.build(story, onFirstPage=self._page_decorator, onLaterPages=self._page_decorator)

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def _build_xlsx(self, result: dict[str, Any], output: Path) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary = result.get("summary", {}) or {}
        observations = summary.get("observations", {}) or {}
        stages = list(result.get("stage_outcomes", []) or [])
        findings = self._collect_findings(stages)
        status_counts = self._status_counts(stages)
        severity_counts = self._severity_counts(findings)

        summary_sheet.append(["TGPAssure Land Gravity QC Report", ""])
        summary_sheet.append(["Metric", "Value"])
        summary_rows = [
            ("Run UUID", result.get("run_uuid")),
            ("Profile", result.get("profile_name")),
            ("Status", str(result.get("status", "")).upper()),
            ("Score", result.get("score")),
            ("Started", result.get("started_at")),
            ("Completed", result.get("completed_at")),
            ("Source Path", observations.get("source_path")),
            ("Survey Type", observations.get("survey_type")),
            ("Record Count", observations.get("record_count")),
            ("Station Count", observations.get("station_count")),
            ("Line Count", observations.get("line_count")),
            ("CRS", observations.get("crs")),
            ("Gravity Units", observations.get("gravity_units")),
            ("Start Time", observations.get("start_time")),
            ("End Time", observations.get("end_time")),
            ("Final Channels", ", ".join(summary.get("final_channels", []) or [])),
        ]
        for row in summary_rows:
            summary_sheet.append(list(row))

        summary_sheet.append([])
        summary_sheet.append(["Stage Status", "Count"])
        status_start = summary_sheet.max_row + 1
        for key in self.STATUS_ORDER:
            if status_counts.get(key, 0):
                summary_sheet.append([key, status_counts[key]])
        status_end = summary_sheet.max_row

        summary_sheet.append([])
        summary_sheet.append(["Finding Severity", "Count"])
        severity_start = summary_sheet.max_row + 1
        for key, value in severity_counts.items():
            if value:
                summary_sheet.append([key, value])
        severity_end = summary_sheet.max_row

        stages_sheet = workbook.create_sheet("QC Stages")
        stages_sheet.append(["Order", "Stage Key", "Stage", "Status", "Score", "Duration ms", "Finding Count", "Message", "Metrics"])
        findings_sheet = workbook.create_sheet("Findings")
        findings_sheet.append(["Severity", "Stage", "Rule", "Finding", "Location", "Recommended Action", "Metadata"])
        metrics_sheet = workbook.create_sheet("Detailed Metrics")
        metrics_sheet.append(["Stage", "Stage Key", "Metric", "Value"])

        for order, stage in enumerate(stages, start=1):
            metrics = stage.get("metrics", {}) or {}
            stages_sheet.append([
                order,
                stage.get("stage_key"),
                stage.get("display_name"),
                str(stage.get("status", "")).upper(),
                stage.get("score", metrics.get("score", metrics.get("qc_score")) if isinstance(metrics, dict) else None),
                stage.get("duration_ms"),
                len(stage.get("findings", []) or []),
                stage.get("message"),
                json.dumps(metrics, ensure_ascii=False, default=str),
            ])
            if isinstance(metrics, dict):
                for key, value in sorted(metrics.items()):
                    metrics_sheet.append([
                        stage.get("display_name"),
                        stage.get("stage_key"),
                        self._humanize(key),
                        self._format_value(value),
                    ])
            for item in stage.get("findings", []) or []:
                findings_sheet.append([
                    str(item.get("severity", "")).upper(),
                    stage.get("display_name"),
                    item.get("rule_id"),
                    item.get("message"),
                    item.get("location_ref"),
                    item.get("suggested_action"),
                    json.dumps(item.get("metadata", {}), ensure_ascii=False, default=str),
                ])

        for title, key in (
            ("Base Statistics", "base_statistics"),
            ("Repeat Stations", "repeat_statistics"),
            ("Loop Closures", "loop_statistics"),
            ("Crossovers", "crossover_statistics"),
            ("Reduction Statistics", "reduction_statistics"),
            ("Anomaly Statistics", "anomaly_statistics"),
        ):
            self._write_data_sheet(workbook.create_sheet(title), summary.get(key) or {})

        thresholds = self._extract_thresholds(result)
        threshold_sheet = workbook.create_sheet("Thresholds")
        threshold_sheet.append(["Threshold", "Configured Value"])
        for key, value in sorted(thresholds.items()):
            threshold_sheet.append([self._humanize(key), self._format_value(value)])

        charts_sheet = workbook.create_sheet("Charts")
        charts_sheet.append(["Charts generated from QC summary tables"])
        if status_end >= status_start:
            chart = PieChart()
            chart.title = "QC Stage Status Distribution"
            data_ref = Reference(summary_sheet, min_col=2, min_row=status_start - 1, max_row=status_end)
            cats_ref = Reference(summary_sheet, min_col=1, min_row=status_start, max_row=status_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.dLbls = DataLabelList()
            chart.dLbls.showPercent = True
            chart.dLbls.showVal = True
            chart.height = 10
            chart.width = 14
            charts_sheet.add_chart(chart, "A3")
        if severity_end >= severity_start:
            chart2 = PieChart()
            chart2.title = "Findings by Severity"
            data_ref = Reference(summary_sheet, min_col=2, min_row=severity_start - 1, max_row=severity_end)
            cats_ref = Reference(summary_sheet, min_col=1, min_row=severity_start, max_row=severity_end)
            chart2.add_data(data_ref, titles_from_data=True)
            chart2.set_categories(cats_ref)
            chart2.dLbls = DataLabelList()
            chart2.dLbls.showPercent = True
            chart2.dLbls.showVal = True
            chart2.height = 10
            chart2.width = 14
            charts_sheet.add_chart(chart2, "A22")
        if stages_sheet.max_row >= 2:
            score_chart = BarChart()
            score_chart.type = "bar"
            score_chart.title = "QC Stage Scores"
            score_chart.x_axis.title = "Score"
            score_chart.y_axis.title = "Stage"
            score_chart.add_data(Reference(stages_sheet, min_col=5, min_row=1, max_row=stages_sheet.max_row), titles_from_data=True)
            score_chart.set_categories(Reference(stages_sheet, min_col=3, min_row=2, max_row=stages_sheet.max_row))
            score_chart.dLbls = DataLabelList()
            score_chart.dLbls.showVal = True
            score_chart.height = max(10, min(17, 7 + stages_sheet.max_row * 0.35))
            score_chart.width = 22
            charts_sheet.add_chart(score_chart, "P3")

            duration_chart = BarChart()
            duration_chart.type = "bar"
            duration_chart.title = "QC Stage Execution Duration"
            duration_chart.x_axis.title = "Duration (ms)"
            duration_chart.y_axis.title = "Stage"
            duration_chart.add_data(Reference(stages_sheet, min_col=6, min_row=1, max_row=stages_sheet.max_row), titles_from_data=True)
            duration_chart.set_categories(Reference(stages_sheet, min_col=3, min_row=2, max_row=stages_sheet.max_row))
            duration_chart.dLbls = DataLabelList()
            duration_chart.dLbls.showVal = True
            duration_chart.height = max(10, min(17, 7 + stages_sheet.max_row * 0.35))
            duration_chart.width = 22
            charts_sheet.add_chart(duration_chart, "P22")

        for sheet in workbook.worksheets:
            self._format_sheet(sheet)
        summary_sheet.freeze_panes = "A3"
        workbook.save(output)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _page_decorator(canvas, doc) -> None:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.units import mm

        canvas.saveState()
        width, height = doc.pagesize
        canvas.setStrokeColor(HexColor("#CBD5E0"))
        canvas.line(10 * mm, 9 * mm, width - 10 * mm, 9 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(HexColor("#4A5568"))
        canvas.drawString(10 * mm, 5.5 * mm, "TGPAssure — Land Gravity Quality Control")
        canvas.drawRightString(width - 10 * mm, 5.5 * mm, f"Page {doc.page}")
        canvas.restoreState()

    def _key_value_table(self, rows: list[list[Any]], styles, key_width: float, value_width: float):
        from reportlab.lib import colors
        from reportlab.platypus import Paragraph, Table, TableStyle

        data = [[Paragraph(str(k), styles["GravSmall"]), Paragraph(self._format_value(v), styles["GravSmall"])] for k, v in rows]
        table = Table(data, colWidths=[key_width, value_width])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EEF5")),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#AAB5C0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return table

    def _standard_table(self, rows: list[list[Any]], widths: list[float], styles, repeat_rows: int = 0, font_size: float = 7.5):
        from reportlab.lib import colors
        from reportlab.platypus import Paragraph, Table, TableStyle

        converted: list[list[Any]] = []
        for r_index, row in enumerate(rows):
            converted_row = []
            for cell in row:
                if hasattr(cell, "wrap"):
                    converted_row.append(cell)
                else:
                    converted_row.append(Paragraph(self._format_value(cell), styles["GravTiny"] if font_size < 7 else styles["GravSmall"]))
            converted.append(converted_row)
        table = Table(converted, repeatRows=repeat_rows, colWidths=widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#AAB5C0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return table

    def _acceptance_table(self, status: str, score: float, status_counts: dict[str, int], severity_counts: dict[str, int], styles):
        from reportlab.lib.units import mm
        recommendation = "ACCEPT"
        if status in {"FAIL", "ERROR"} or severity_counts.get("CRITICAL", 0) or severity_counts.get("ERROR", 0):
            recommendation = "HOLD / CORRECT BEFORE ACCEPTANCE"
        elif status == "WARN" or severity_counts.get("WARNING", 0) or severity_counts.get("WARN", 0):
            recommendation = "CONDITIONAL ACCEPTANCE — REVIEW WARNINGS"
        rows = [
            ["Decision", "Score", "Pass", "Warn", "Fail/Error", "Critical/Error findings"],
            [
                recommendation,
                f"{score:.1f}/100",
                status_counts.get("PASS", 0),
                status_counts.get("WARN", 0),
                status_counts.get("FAIL", 0) + status_counts.get("ERROR", 0),
                severity_counts.get("CRITICAL", 0) + severity_counts.get("ERROR", 0),
            ],
        ]
        return self._standard_table(rows, [78 * mm, 32 * mm, 30 * mm, 30 * mm, 40 * mm, 52 * mm], styles, repeat_rows=1)

    def _inventory_rows(self, summary: dict[str, Any], observations: dict[str, Any]) -> list[list[Any]]:
        rows = [
            ["Source file", observations.get("source_path", "")],
            ["Survey type", observations.get("survey_type", "")],
            ["Data role", observations.get("data_role", observations.get("role", ""))],
            ["Instrument", observations.get("instrument", observations.get("instrument_model", ""))],
            ["Serial number", observations.get("instrument_serial", observations.get("serial_number", ""))],
            ["CRS", observations.get("crs", "") or "Not defined"],
            ["Gravity datum", observations.get("gravity_datum", summary.get("gravity_datum", "Not reported"))],
            ["Elevation datum", observations.get("elevation_datum", summary.get("elevation_datum", "Not reported"))],
            ["Reduction density", summary.get("density_g_cm3", observations.get("density_g_cm3", "Not reported"))],
            ["Records / Stations / Lines", f"{self._safe_int(observations.get('record_count')):,} / {self._safe_int(observations.get('station_count')):,} / {self._safe_int(observations.get('line_count')):,}"],
            ["Start / End", f"{observations.get('start_time', '')}  →  {observations.get('end_time', '')}"],
            ["Available channels", ", ".join(observations.get("channels", []) or summary.get("available_channels", []) or []) or "Not reported"],
            ["Final channels", ", ".join(summary.get("final_channels", []) or []) or "Not reported"],
        ]
        return rows

    def _correction_chain_rows(self, summary: dict[str, Any], stages: list[dict[str, Any]]) -> list[list[Any]]:
        keywords = (
            ("Tidal correction", ("tidal", "tide")),
            ("Instrument drift", ("drift",)),
            ("Latitude / normal gravity", ("latitude", "normal_gravity")),
            ("Free-air correction", ("free_air", "free-air")),
            ("Bouguer correction", ("bouguer",)),
            ("Terrain correction", ("terrain",)),
            ("Loop / network adjustment", ("loop", "network")),
            ("Cross-over consistency", ("crossover", "cross-over")),
            ("Final anomaly consistency", ("final_anomaly", "anomaly")),
        )
        rows: list[list[Any]] = []
        for label, keys in keywords:
            matches = []
            for stage in stages:
                text = f"{stage.get('stage_key', '')} {stage.get('display_name', '')}".lower()
                if any(key in text for key in keys):
                    matches.append(stage)
            if not matches:
                continue
            stage = matches[0]
            metric_preview = self._metric_preview(stage.get("metrics", {}) or {})
            rows.append([label, str(stage.get("status", "")).upper(), metric_preview or stage.get("message", "")])
        return rows

    def _dynamic_records_table(self, rows: list[dict[str, Any]], styles):
        keys = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        keys = keys[:8]
        data = [[self._humanize(key) for key in keys]]
        for row in rows[:200]:
            data.append([self._format_value(row.get(key)) for key in keys])
        total_width = 262.0
        width = total_width / max(len(keys), 1)
        from reportlab.lib.units import mm
        return self._standard_table(data, [width * mm] * len(keys), styles, repeat_rows=1, font_size=6.7)

    def _plot_status_counts(self, counts: dict[str, int], Image):
        labels = [key for key in self.STATUS_ORDER if counts.get(key, 0)]
        values = [counts[key] for key in labels]
        return self._matplotlib_bar("QC Stage Status Distribution", labels, values, "Stages", Image)

    def _plot_severity_counts(self, counts: dict[str, int], Image):
        labels = [key for key, value in counts.items() if value]
        values = [counts[key] for key in labels]
        return self._matplotlib_bar("Findings by Severity", labels, values, "Findings", Image)

    def _plot_stage_durations(self, stages: list[dict[str, Any]], Image):
        rows = [(str(stage.get("display_name") or stage.get("stage_key") or "Stage"), self._safe_float(stage.get("duration_ms"), 0.0)) for stage in stages]
        rows = [(label, value) for label, value in rows if value > 0]
        if not rows:
            return None
        rows = sorted(rows, key=lambda item: item[1], reverse=True)[:15]
        return self._matplotlib_bar("Longest QC Stages", [r[0] for r in rows], [r[1] for r in rows], "Duration (ms)", Image, horizontal=True)

    def _plot_stage_metric_score(self, stages: list[dict[str, Any]], Image):
        labels: list[str] = []
        values: list[float] = []
        for stage in stages:
            metrics = stage.get("metrics", {}) or {}
            candidate = stage.get("score")
            if candidate is None and isinstance(metrics, dict):
                candidate = metrics.get("score", metrics.get("qc_score"))
            if candidate is None:
                continue
            value = self._safe_float(candidate, math.nan)
            if math.isfinite(value):
                labels.append(str(stage.get("display_name") or stage.get("stage_key") or "Stage"))
                values.append(value)
        if not values:
            return None
        return self._matplotlib_bar("QC Stage Scores", labels[:20], values[:20], "Score", Image, horizontal=True)

    def _plot_record_values(self, title: str, data: Any, Image):
        rows = self._records_to_rows(data)
        if len(rows) < 2:
            return None
        numeric_keys: list[str] = []
        for key in rows[0].keys():
            numeric_count = sum(self._is_number(row.get(key)) for row in rows)
            if numeric_count >= min(3, len(rows)):
                numeric_keys.append(key)
        if not numeric_keys:
            return None
        value_key = numeric_keys[-1]
        labels = [str(row.get("line_id") or row.get("station_id") or row.get("loop_id") or row.get("id") or index + 1) for index, row in enumerate(rows[:30])]
        values = [self._safe_float(row.get(value_key), 0.0) for row in rows[:30]]
        return self._matplotlib_bar(f"{title}: {self._humanize(value_key)}", labels, values, self._humanize(value_key), Image, horizontal=len(labels) > 10)

    @staticmethod
    def _matplotlib_bar(title: str, labels: list[str], values: list[float], ylabel: str, Image, horizontal: bool = False):
        if not labels or not values:
            return None
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
        except Exception:
            return None
        fig, ax = plt.subplots(figsize=(9.2, 4.2))
        if horizontal:
            positions = list(range(len(labels)))
            ax.barh(positions, values)
            ax.set_yticks(positions)
            ax.set_yticklabels([label[:42] for label in labels], fontsize=7)
            ax.invert_yaxis()
            ax.set_xlabel(ylabel)
        else:
            positions = list(range(len(labels)))
            ax.bar(positions, values)
            ax.set_xticks(positions)
            ax.set_xticklabels([label[:24] for label in labels], rotation=30, ha="right", fontsize=7)
            ax.set_ylabel(ylabel)
        ax.set_title(title)
        ax.grid(axis="y" if not horizontal else "x", alpha=0.25)
        fig.tight_layout()
        buffer = io.BytesIO()
        fig.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
        plt.close(fig)
        buffer.seek(0)
        from reportlab.lib.units import mm
        image = Image(buffer, width=126 * mm, height=58 * mm)
        image._tgp_buffer = buffer
        return image

    @staticmethod
    def _collect_findings(stages: list[dict[str, Any]]) -> list[tuple[str, dict[str, Any]]]:
        findings: list[tuple[str, dict[str, Any]]] = []
        for stage in stages:
            name = stage.get("display_name") or stage.get("stage_key") or "Unknown"
            findings.extend((str(name), item) for item in (stage.get("findings", []) or []))
        return findings

    @classmethod
    def _status_counts(cls, stages: list[dict[str, Any]]) -> dict[str, int]:
        counts = {key: 0 for key in cls.STATUS_ORDER}
        for stage in stages:
            key = str(stage.get("status", "UNKNOWN")).upper()
            counts[key] = counts.get(key, 0) + 1
        return counts

    @classmethod
    def _severity_counts(cls, findings: list[tuple[str, dict[str, Any]]]) -> dict[str, int]:
        counts: dict[str, int] = {}
        for _, finding in findings:
            key = str(finding.get("severity", "INFO")).upper()
            counts[key] = counts.get(key, 0) + 1
        return counts

    def _executive_summary_text(self, status: str, score: float, status_counts: dict[str, int], severity_counts: dict[str, int], findings: list[tuple[str, dict[str, Any]]]) -> str:
        critical = severity_counts.get("CRITICAL", 0) + severity_counts.get("ERROR", 0)
        warnings = severity_counts.get("WARNING", 0) + severity_counts.get("WARN", 0)
        fail_stages = status_counts.get("FAIL", 0) + status_counts.get("ERROR", 0)
        if status in {"FAIL", "ERROR"} or critical or fail_stages:
            decision = "The dataset should be held for corrective review before final gravity-data acceptance."
        elif status == "WARN" or warnings:
            decision = "The dataset is conditionally acceptable subject to review of the flagged warnings and supporting field/reduction records."
        else:
            decision = "No material automated QC failure was detected; the dataset is suitable for technical review toward acceptance."
        return (
            f"Overall automated result: <b>{status}</b> with a score of <b>{score:.1f}/100</b>. "
            f"The pipeline recorded {status_counts.get('PASS', 0)} passing stages, {status_counts.get('WARN', 0)} warning stages and {fail_stages} failed/error stages. "
            f"There are {critical} critical/error findings and {warnings} warning findings. {decision}"
        )

    def _recommendations(self, findings: list[tuple[str, dict[str, Any]]], status: str) -> list[str]:
        actions: list[str] = []
        for _, item in findings:
            action = str(item.get("suggested_action", "") or "").strip()
            if action and action not in actions:
                actions.append(action)
            if len(actions) >= 12:
                break
        if not actions:
            if status in {"FAIL", "ERROR"}:
                actions.append("Review failed QC stages, field logs, base-loop records and all applied gravity corrections before release.")
            elif status == "WARN":
                actions.append("Review all warning findings and confirm that observed deviations are explained by field conditions or geology.")
            else:
                actions.append("Retain the QC report with the processed deliverables and obtain final geophysicist sign-off.")
        return actions

    def _extract_thresholds(self, result: dict[str, Any]) -> dict[str, Any]:
        candidates = [
            result.get("thresholds"),
            (result.get("parameters") or {}).get("thresholds") if isinstance(result.get("parameters"), dict) else None,
            (result.get("summary") or {}).get("thresholds") if isinstance(result.get("summary"), dict) else None,
            result.get("profile_thresholds"),
        ]
        for candidate in candidates:
            if isinstance(candidate, dict) and candidate:
                return candidate
        return {}

    @staticmethod
    def _records_to_rows(data: Any) -> list[dict[str, Any]]:
        if isinstance(data, list):
            return [row for row in data if isinstance(row, dict)]
        if isinstance(data, dict):
            if all(isinstance(value, dict) for value in data.values()) and data:
                rows = []
                for key, value in data.items():
                    row = {"id": key}
                    row.update(value)
                    rows.append(row)
                return rows
            return [{"metric": key, "value": value} for key, value in data.items()]
        return []

    @staticmethod
    def _metric_preview(metrics: dict[str, Any], limit: int = 4) -> str:
        if not isinstance(metrics, dict):
            return ""
        parts = []
        for key, value in metrics.items():
            if isinstance(value, (dict, list, tuple)):
                continue
            parts.append(f"{GravityQcReportBuilder._humanize(key)}: {value}")
            if len(parts) >= limit:
                break
        return "; ".join(parts)

    @staticmethod
    def _format_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if math.isnan(value):
                return "NaN"
            if abs(value) >= 1000:
                return f"{value:,.3f}"
            return f"{value:.6g}"
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return str(value)

    @staticmethod
    def _humanize(value: Any) -> str:
        return str(value).replace("_", " ").strip().title()

    @staticmethod
    def _safe_float(value: Any, default: float = 0.0) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _safe_int(value: Any, default: int = 0) -> int:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _is_number(value: Any) -> bool:
        try:
            float(value)
            return value is not None
        except (TypeError, ValueError):
            return False

    def _write_data_sheet(self, sheet, data: Any) -> None:
        rows = self._records_to_rows(data)
        if not rows:
            sheet.append(["No data"])
            return
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        sheet.append([self._humanize(key) for key in keys])
        for row in rows:
            sheet.append([self._format_value(row.get(key)) for key in keys])

    @staticmethod
    def _format_sheet(sheet) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if sheet.max_row:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2C3E50")
        if sheet.max_row >= 2 and sheet.title != "Summary":
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            width = min(
                max(
                    max((len(str(cell.value or "")) for cell in sheet[get_column_letter(column)]), default=10) + 2,
                    12,
                ),
                60,
            )
            sheet.column_dimensions[get_column_letter(column)].width = width
